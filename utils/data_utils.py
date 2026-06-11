"""
utils/data_utils.py — TIP ESG Platform · Data Write Helpers
============================================================
Functions that save/load supplementary data and build master CSV rows.
Extracted from app.py to allow page modules to import without
circular dependency.

Note: _save_submission_to_csv updates in-memory globals in app.py via
state module after writing to disk.
"""
from __future__ import annotations
import pandas as pd
import numpy as np
import logging
from pathlib import Path
from datetime import datetime
from filelock import FileLock

import config as cfg
import data_loader as dl
import state

_log = logging.getLogger("esg_app")

def _load_supplementary(company: str, year: int) -> dict:
    """Load supplementary fields for company+year. Returns {} if not found."""
    from pathlib import Path as _P
    import csv
    p = _P("data_storage/master/ESG_SUPPLEMENTARY.csv")
    if not p.exists(): return {}
    with open(p, newline="", encoding="utf-8") as f:
        for row in csv.DictReader(f):
            if row.get("Company","").strip()==company and str(row.get("Year","")).strip()==str(year):
                return {k: float(v) if v else 0.0 for k,v in row.items()
                        if k not in ("Company","Year")}
    return {}


def _save_supplementary(company: str, year: int, data: dict) -> None:
    """Upsert supplementary fields for company+year."""
    from pathlib import Path as _P
    import csv
    p = _P("data_storage/master/ESG_SUPPLEMENTARY.csv")
    p.parent.mkdir(parents=True, exist_ok=True)
    rows = []
    if p.exists():
        with open(p, newline="", encoding="utf-8") as f:
            for row in csv.DictReader(f):
                if not (row.get("Company","").strip()==company and
                        str(row.get("Year","")).strip()==str(year)):
                    rows.append(row)
    new_row = {"Company": company, "Year": str(year)}
    for field in state._SUPP_FIELDS[2:]:          # skip Company, Year
        new_row[field] = str(data.get(field, 0) or 0)
    rows.append(new_row)
    with open(p, "w", newline="", encoding="utf-8") as f:
        w = csv.DictWriter(f, fieldnames=state._SUPP_FIELDS)
        w.writeheader(); w.writerows(rows)


def _build_master_row(inp, out, supp: dict = None) -> dict:
    """
    Build a dict whose keys exactly match the master wide CSV column names.
    Ensures no duplicate columns when appended to the master CSV.
    Includes waste KPIs and electricity-by-country columns.
    """
    renew_share  = (inp.renew_elec_purchased + inp.self_gen_elec) / max(out.total_electricity, 1) * 100
    scope1_share = out.total_co2_scope1 / max(out.total_co2, 1) * 100
    scope2_share = out.total_co2_scope2 / max(out.total_co2, 1) * 100
    fuel_total   = (inp.nat_gas + inp.coal_sub + inp.propane + inp.fuel_oil_heavy_a + inp.diesel + inp.petrol)
    fossil_share = fuel_total / max(out.total_energy, 1) * 100
    prod         = max(inp.production, 1)

    # Waste derived
    waste_total    = float(getattr(inp, "waste_total", 0) or 0)
    waste_recovery = float(getattr(inp, "waste_recovery", 0) or 0)
    recovery_rate  = round(waste_recovery / waste_total * 100, 4) if waste_total else 0.0
    waste_elim     = round(waste_total - waste_recovery, 4)

    row = {
        "Company": inp.company, "Year": inp.year,
        "Total no. of sites": int(round(inp.total_sites)),
        "ISO 14001 sites":    int(round(inp.iso_sites)),
        "% certified sites":  round(out.pct_certified, 6),
        "Production":         round(inp.production, 4),
        "Water intake":       round(inp.water_withdrawals, 4),
        "Water intake - KPI": round(out.water_kpi, 6),
        "Total Electricity":                               round(out.total_electricity, 4),
        "Renewable Electricity Purchased":                 round(inp.renew_elec_purchased, 4),
        "Non-Renewable Electricity Purchased":             round(inp.nonrenew_elec_purchased, 4),
        "Self-generated AND consumed electricity on-site": round(inp.self_gen_elec, 4),
        "Purchased Steam":   round(inp.purchased_steam, 4),
        "Sold Electricity":  round(inp.sold_electricity, 4),
        "Sold Steam":        round(inp.sold_steam, 4),
        "Natural Gas":       round(inp.nat_gas, 4),
        "Coal":              round(inp.coal_sub, 4),
        "Propane":           round(inp.propane, 4),
        "Fuel Oil":          round(inp.fuel_oil_heavy_a, 4),
        "Diesel":            round(inp.diesel, 4),
        "Petrol":            round(inp.petrol, 4),
        "Biomass":           round(inp.biomass, 4),
        "Waste tires":       round(inp.waste_tires_mt, 4),
        "LPG":               round(inp.lpg, 4),
        "Other":             round(inp.other_fuels, 4),
        "Total energy":          round(out.total_energy, 4),
        "Total energy - KPI":    round(out.energy_kpi, 6),
        "Total CO2 - Scope 1":   round(out.total_co2_scope1, 4),
        "Total CO2 - Scope 2":   round(out.total_co2_scope2, 4),
        "Total CO2":             round(out.total_co2, 4),
        "Total CO2 - KPI":       round(out.co2_kpi, 6),
        # ── Waste fields ──────────────────────────────────────────────────────
        "Total Waste":           round(waste_total, 4),
        "Waste Recovered":       round(waste_recovery, 4),
        "Recovery Rate":         recovery_rate,
        # ── Country electricity placeholders (filled by _save_electricity_to_master) ─
        **{_elec_col(c): None for c in state.ELEC_ALL_COUNTRIES},
        # ── Derived KPIs ──────────────────────────────────────────────────────
        "Renewable_Electricity_Share_%": round(renew_share, 4),
        "Scope1_Share_%":                round(scope1_share, 4),
        "Scope2_Share_%":                round(scope2_share, 4),
        "Fossil_Energy_Share_%":         round(fossil_share, 4),
        "Water_per_ton":                 round(inp.water_withdrawals / prod, 4),
        "CO2_per_ton":                   round(out.total_co2 / prod, 4),
        "Energy_per_ton":                round(out.total_energy / prod, 4),
        "ISO_Certification_%":           round(out.pct_certified * 100, 2),
        "Waste_Recovery_Rate_%":         recovery_rate,
        "Total_Electricity_by_Country_GJ": None,  # filled after country save
    }

    # ── People & Governance fields (from supplementary) ────────────────────
    s = supp or {}
    def _sf(k, default=0.0):
        try: return float(s.get(k, default) or default)
        except: return default

    stress_wd     = _sf("stress_water_withdrawal")
    non_stress_wd = _sf("non_stress_water_withdrawal",
                         max(inp.water_withdrawals - stress_wd, 0))
    hs_ext   = _sf("hs_external_audit")
    hs_int   = _sf("hs_internal_audit")
    hs_tot   = max(int(_sf("hs_external_audit", inp.total_sites)), int(inp.total_sites))
    emp_tot  = _sf("total_employees")
    emp_fem  = _sf("female_employees")
    bod_tot  = _sf("board_total")
    bod_fem  = _sf("female_board")

    row.update({
        # Water detail
        "Stress Water Withdrawal":     round(stress_wd, 4),
        "Non-Stress Water Withdrawal": round(non_stress_wd, 4),
        # Coal breakdown
        "Coal Sub-Bituminous":         round(_sf("coal_sub_bituminous"), 4),
        "Coal Brown Briquettes":       round(_sf("coal_brown_briquettes"), 4),
        "Coal Other Bituminous":       round(_sf("coal_other_bituminous"), 4),
        # H&S
        "HS External Audit Sites":     round(hs_ext),
        "HS Internal Audit Sites":     round(hs_int),
        "HS External Audit %":         round(hs_ext / max(hs_tot, 1) * 100, 2),
        "HS Internal Audit %":         round(hs_int / max(hs_tot, 1) * 100, 2),
        # Diversity
        "Total Employees":             round(emp_tot),
        "Female Employees":            round(emp_fem),
        "Female Employees %":          round(emp_fem / max(emp_tot, 1) * 100, 2),
        "Board Total":                 round(bod_tot),
        "Female Board":                round(bod_fem),
        "Female Board %":              round(bod_fem / max(bod_tot, 1) * 100, 2),
        # Science-Based Targets
        "SBT Total":       round(_sf("sbt_total")),
        "SBT Validated":   round(_sf("sbt_validated")),
        "SBT Committed":   round(_sf("sbt_committed")),
        "SBT Non-Committed": round(_sf("sbt_non_committed")),
    })
    return row


def _save_version_parquet(inp, combined_df: pd.DataFrame) -> str:
    """
    Save the ENTIRE company template (all years) as a Parquet snapshot.
    Stored in data_storage/versions/{CompanyName}/ — subfolder only, never flat.
    Filename: CompanyName_Year_YYYYMMDD_HHMMSS.parquet (year = the year just edited).
    NEVER overwritten — each save event creates a new file (full audit trail).
    Reading this file shows the complete state of all years for that company
    at the exact moment the save was made.
    """
    from pathlib import Path
    from datetime import datetime
    ts      = datetime.now().strftime("%Y%m%d_%H%M%S")
    co_safe = inp.company.replace(" ", "_").replace("/", "_")
    # Extract ALL rows for this company from the combined master DataFrame
    company_all_years = combined_df[combined_df["Company"] == inp.company].copy()
    filename = f"{co_safe}_{inp.year}_{ts}.parquet"
    # Subfolder only — no flat file
    ver_dir  = Path("data_storage") / "versions" / co_safe
    ver_dir.mkdir(parents=True, exist_ok=True)
    try:
        company_all_years.to_parquet(ver_dir / filename, index=False)
        return f"{co_safe}/{filename}"
    except Exception as e:
        return f"[version save failed: {e}]"


def _drop_zero_elec_cols(df: "pd.DataFrame") -> "pd.DataFrame":
    """
    Return df with Elec_*_GJ country columns removed if every value in that
    column is zero or null across all rows.  Non-electricity columns are
    never touched.  Used so files only carry countries with actual consumption.
    """
    elec_cols = [c for c in df.columns if c.startswith("Elec_") and c.endswith("_GJ")
                 and c != "Total_Electricity_by_Country_GJ"]
    zero_cols = [c for c in elec_cols
                 if df[c].fillna(0).eq(0).all()]
    return df.drop(columns=zero_cols) if zero_cols else df


def _sync_company_member_files(master_df: "pd.DataFrame") -> list:
    """
    Write per-company CSVs in data_storage/members/TIP/<CompanyName>/<CompanyName>_latest.csv
    from the current master wide DataFrame.
    Skips any file that is locked (e.g. open in Excel) instead of crashing.
    Returns list of company names that were skipped.
    """
    from pathlib import Path
    members_tip = Path("data_storage/members/TIP")
    skipped = []
    for company, grp in master_df.groupby("Company"):
        co_safe   = str(company).replace(" ", "_")
        co_folder = members_tip / co_safe
        co_folder.mkdir(parents=True, exist_ok=True)
        try:
            # Drop electricity country columns that are all zero for this company
            grp_clean = _drop_zero_elec_cols(grp.reset_index(drop=True))
            grp_clean.to_csv(co_folder / f"{co_safe}_latest.csv", index=False)
        except (PermissionError, OSError):
            skipped.append(str(company))
    return skipped


def _update_tip_members_file(master_path: "Path", tip_master_path: "Path") -> None:
    """Rebuild the TIP members aggregate strictly from the latest master on disk.

    This prevents mismatches where the in-memory combined_df used during save
    (bootstrap/reconstruction) differs from the finally-written master CSV.
    """
    import pandas as pd
    tip_master_path.parent.mkdir(parents=True, exist_ok=True)
    try:
        master_df = pd.read_csv(master_path)
    except Exception as e:
        _log.error("[tip_members] Could not read master to rebuild tip members: %s", e)
        return

    try:
        _drop_zero_elec_cols(master_df).to_csv(tip_master_path, index=False)
    except Exception as e:
        _log.error("[tip_members] Could not write TIP members file: %s", e)


def _migrate_supplementary_to_master() -> str:
    """
    One-time (idempotent) migration: read every row in ESG_SUPPLEMENTARY.csv
    and upsert it into the master CSV so the People & Governance columns are
    populated for all previously-submitted records.
    Safe to call on every startup — skips companies/years already promoted.
    """
    import csv as _csv
    from pathlib import Path as _P
    supp_path = _P("data_storage/master/ESG_SUPPLEMENTARY.csv")
    if not supp_path.exists():
        return "no supplementary file — skipped"

    migrated = 0
    with open(supp_path, newline="", encoding="utf-8") as f:
        rows = list(_csv.DictReader(f))

    for row in rows:
        company = row.get("Company","").strip()
        year_s  = row.get("Year","").strip()
        if not company or not year_s:
            continue
        try:
            year = int(year_s)
        except ValueError:
            continue

        # Build supp dict for this row
        supp = {k: float(v) if v else 0.0
                for k, v in row.items() if k not in ("Company","Year")}

        # Load existing TemplateInputs for this company+year
        hist = dl.get_company_hist(state.CONSOLIDATED_DF, company)
        if not hist:
            continue
        sd = dl.get_step_data(hist, year)
        sd_clean = {k: v for k, v in sd.items() if k in state.VALID_TEMPLATE_FIELDS}
        if not sd_clean:
            continue

        inp = TemplateInputs(company=company, year=year, **sd_clean)
        out = calculate(inp)

        # Re-save the master row with supplementary data included
        _save_submission_to_csv(inp, out)   # supp auto-loaded inside
        migrated += 1

    return f"migrated {migrated} supplementary records into master"


def _write_verification_status(company: str, year: int, status: str) -> None:
    """
    Persist DSS+ verification status for a company+year to a CSV file.
    Status values: 'Verified', 'Pending', 'Flagged'.
    Client home page reads this file to show the verification chip.
    """
    from pathlib import Path
    import csv, os

    vcsv = Path("data_storage/verifications.csv")
    vcsv.parent.mkdir(parents=True, exist_ok=True)

    rows = []
    if vcsv.exists():
        with open(vcsv, newline="") as f:
            for row in csv.DictReader(f):
                if not (row.get("Company","").strip() == company and
                        str(row.get("Year","")).strip() == str(year)):
                    rows.append(row)   # keep other company/year rows

    rows.append({"Company": company, "Year": str(year), "Status": status,
                 "UpdatedBy": "dss+ Analyst"})

    with open(vcsv, "w", newline="") as f:
        w = csv.DictWriter(f, fieldnames=["Company","Year","Status","UpdatedBy"])
        w.writeheader()
        w.writerows(rows)


def _save_submission_to_csv(inp, out) -> str:
    """
    Three independent operations:

    1. MASTER CSV (data_storage/master/) — overwrite the row for this company+year.
       The master always holds the LATEST values. Second save for same company+year
       replaces the first row.

    2. VERSION Parquet (data_storage/versions/) — always ADD a new timestamped file.
       Never overwritten. Full audit trail of every save event.

    3. SYNC (after master is written):
       - CONSOLIDATED_DUMMY Excel Raw Dummy data sheet (long format)
       - Per-company CSVs in data_storage/members/TIP/<Company>/
       - TIP members aggregate CSV
    """
    import os, tempfile
    from pathlib import Path
    from datetime import datetime

    _master_cands = dl._get_csv_candidates()
    csv_path = next((p for p in _master_cands if p.exists()
                     and p.name.startswith("ESG_MASTER_WIDE_ALL_COMPANIES_")),
                    None) or Path("data_storage/master/ESG_MASTER_WIDE_ALL_COMPANIES_2009_2023.csv")
    csv_path.parent.mkdir(parents=True, exist_ok=True)

    # Load supplementary fields so they are promoted into the master CSV
    _supp_data = _load_supplementary(inp.company, inp.year)
    new_row    = pd.DataFrame([_build_master_row(inp, out, supp=_supp_data)])
    master_cols = list(new_row.columns)

    def _align(df):
        """Align DataFrame to master column order: strip extras, fill missing."""
        if df.empty:
            return pd.DataFrame(columns=master_cols)
        extra = [c for c in df.columns if c not in master_cols]
        if extra:
            df = df.drop(columns=extra)
        for col in master_cols:
            if col not in df.columns:
                df[col] = None
        return df[master_cols]

    def _load_best_existing():
        """
        Load the most complete existing master DataFrame.
        Checks all candidate paths and picks the one with the most rows.
        """
        candidates = [
            csv_path,
            Path("data_storage/raw/ESG_MASTER_WIDE_ALL_COMPANIES_2009_2023.csv"),
        ]
        best = pd.DataFrame(columns=master_cols)
        for p in candidates:
            if p.exists():
                try:
                    df = pd.read_csv(p)
                    if "Company" in df.columns and "Year" in df.columns and len(df) > len(best):
                        best = df
                except PermissionError:
                    pass
                except Exception:
                    pass
        return _align(best)

    # ── 1. Build combined DataFrame ──────────────────────────────────────────
    # H1 FIX: advisory file lock held for the full read-modify-write cycle.
    # Prevents data corruption if two analysts save simultaneously.
    # Timeout=10s: if another process holds the lock and crashes, we don't
    # block forever.  The PermissionError branch below still handles Excel locks.
    lock_path = csv_path.with_suffix(".lock")
    with FileLock(str(lock_path), timeout=cfg.FILELOCK_TIMEOUT):
        existing = _load_best_existing()
        mask     = ~((existing["Company"] == inp.company) & (existing["Year"] == inp.year))
        existing = existing[mask]
        combined = pd.concat([existing, new_row], ignore_index=True)
        combined = combined.sort_values(["Company", "Year"]).reset_index(drop=True)

        n_records   = len(combined)
        n_companies = combined["Company"].nunique()

        # ── 2. Save version Parquet BEFORE touching master (audit trail first) ───
        version_filename = _save_version_parquet(inp, combined)

        # ── 3. Write master CSV, then sync all dependent files ───────────────────
        try:
            # Master CSV keeps all country columns (even all-zero) as the full schema.
            # Derived outputs (member files, TIP aggregate) strip all-zero country cols.
            combined.to_csv(csv_path, index=False)
            # Sync TIP members aggregate
            tip_master_path = Path("data_storage/members/TIP/ESG_MASTER_WIDE_TIP_MEMBERS_2009_2023.csv")
            _update_tip_members_file(csv_path, tip_master_path)
            # Sync per-company member files
            _sync_company_member_files(combined)
            # Sync CONSOLIDATED_DUMMY Excel
            _sync_consolidate_excel(combined)

            # ── Auto-add electricity-by-country year columns for new submission ──
            # If company just submitted for a year that isn't in the electricity
            # editor yet, initialize all country columns to 0 in the master CSV.
            _new_yr = inp.year
            _elec_cols_all = [c for c in combined.columns
                              if c.startswith("Elec_") and c.endswith("_GJ")]
            _co_yr_mask = (combined["Company"] == inp.company) & (combined["Year"] == _new_yr)
            if _co_yr_mask.any() and _elec_cols_all:
                for _ec in _elec_cols_all:
                    if pd.isna(combined.loc[_co_yr_mask, _ec]).all():
                        combined.loc[_co_yr_mask, _ec] = 0.0
                # Also ensure Total_Electricity_by_Country_GJ exists
                if "Total_Electricity_by_Country_GJ" not in combined.columns:
                    combined["Total_Electricity_by_Country_GJ"] = 0.0
                elif pd.isna(combined.loc[_co_yr_mask, "Total_Electricity_by_Country_GJ"]).all():
                    combined.loc[_co_yr_mask, "Total_Electricity_by_Country_GJ"] = 0.0
                # Re-write master with the zero-initialized electricity columns
                combined.to_csv(csv_path, index=False)

            # ── CRITICAL: update the in-memory globals so all pages in this
            #    session immediately see the new data without requiring a restart.
            state.CONSOLIDATED_DF     = combined.copy()
            state.COMPANIES           = dl.get_companies(combined)
            state.USING_FALLBACK = False
            try:
                cfg.refresh_year_bounds(combined)
                state.HIST_YEARS = cfg.hist_years()
                state.CURR_YEAR  = cfg.curr_year()
                state.LONG_YEARS = cfg.long_years()
            except Exception:
                pass
            try:
                state.SECTOR_DF = dl.load_sector_aggregated(combined)
            except Exception:
                pass
            # LONG_DATA/FUEL_MIX are rebuilt by app.py after state update
            try:
                import streamlit as _st; _st.cache_data.clear()
            except Exception: pass

            return (f"Saved {inp.company} — {inp.year}. "
                    f"Master: {n_records} records across {n_companies} companies. "
                    f"Version: {version_filename}")
        except PermissionError:
            ts          = datetime.now().strftime("%Y%m%d_%H%M%S")
            backup_name = f"ESG_MASTER_{inp.company.replace(' ','_')}_{inp.year}_{ts}.csv"
            backup_path = csv_path.parent / backup_name
            try:
                combined.to_csv(backup_path, index=False)
                return (
                    f"⚠️ Master file open in Excel — saved backup: **{backup_name}**\n"
                    f"Version snapshot: {version_filename}\n"
                    f"Close Excel and click Save again."
                )
            except Exception as e2:
                return f"❌ Save failed (file locked AND backup failed): {e2}"
        except Exception as e:
            return f"❌ Save failed: {e}"


def _save_electricity_to_master(company: str, year: int) -> str:
    """
    Save electricity-by-country data (from the Electricity tab editor) into:
      1. Master wide CSV  — columns Elec_<Country>_GJ  (GJ = MWh x 3.6)
      2. TIP members aggregate CSV
      3. Per-company member CSVs in data_storage/members/TIP/<Company>/
      4. CONSOLIDATED_DUMMY Excel (Raw Dummy data sheet, long format)
      5. Parquet snapshot of the complete company+year row

    Updates ALL years that have non-zero values in the electricity editor.
    Only the 7 countries already in the master schema are written:
        Canada, Mexico, United States, Japan, France, Hungary, Italy
    Any other country rows in the editor UI are displayed but not persisted.
    """
    from pathlib import Path
    from datetime import datetime

    COUNTRY_COL = state.ELEC_COUNTRY_COLS  # all 31 countries
    MWH_TO_GJ = 3.6

    elec_df = st.session_state.get("elec_data", pd.DataFrame())
    if elec_df.empty:
        return "No electricity data entered yet."

    _ecands = dl._get_csv_candidates()
    csv_path = next((p for p in _ecands if p.exists()
                     and p.name.startswith("ESG_MASTER_WIDE_ALL_COMPANIES_")), None)
    if csv_path is None:
        return "Master CSV not found. Save KPI data first."
    try:
        master = pd.read_csv(csv_path)
    except PermissionError:
        return "Master CSV is open in Excel — close it and try again."

    # Ensure all country columns exist in master (add if missing)
    for col in COUNTRY_COL.values():
        if col not in master.columns:
            master[col] = None
    if "Total_Electricity_by_Country_GJ" not in master.columns:
        master["Total_Electricity_by_Country_GJ"] = None

    yr_cols = [c for c in elec_df.columns if str(c).isdigit() and 2000 < int(c) < 2030]

    updated_years = []
    for yr_str in yr_cols:
        yr   = int(yr_str)
        mask = (master["Company"] == company) & (master["Year"] == yr)
        if not mask.any():
            # KPI row not found for this year — create a minimal stub row so
            # electricity data is not lost; user can submit KPIs later.
            stub = pd.DataFrame([{
                "Company": company, "Year": yr,
                **{c: 0.0 for c in COUNTRY_COL.values()},
                "Total_Electricity_by_Country_GJ": 0.0,
            }])
            # Align to master columns
            for col in master.columns:
                if col not in stub.columns:
                    stub[col] = None
            stub = stub[master.columns]
            master = pd.concat([master, stub], ignore_index=True)
            master = master.sort_values(["Company", "Year"]).reset_index(drop=True)
            mask = (master["Company"] == company) & (master["Year"] == yr)

        year_series = elec_df.set_index("Country")[yr_str]
        for country, mwh_val in year_series.items():
            col_name = COUNTRY_COL.get(str(country))
            if col_name is None:
                continue  # country not in master schema
            gj_val = float(mwh_val) * MWH_TO_GJ if pd.notna(mwh_val) else 0.0
            master.loc[mask, col_name] = round(gj_val, 4)

        # Recompute total-by-country for this row
        country_vals = [master.loc[mask, c].values[0]
                        for c in COUNTRY_COL.values() if c in master.columns]
        master.loc[mask, "Total_Electricity_by_Country_GJ"] = round(
            sum(v for v in country_vals if pd.notna(v)), 4)

        updated_years.append(yr)

    if not updated_years:
        return f"No KPI rows found for {company}. Save KPI data first."

    try:
        # H1 FIX: use the same advisory lock as the KPI save path
        lock_path = csv_path.with_suffix(".lock")
        with FileLock(str(lock_path), timeout=cfg.FILELOCK_TIMEOUT):
            master.to_csv(csv_path, index=False)
    except PermissionError:
        return "Master CSV is open in Excel — close it and try again."

    # Sync all dependent files
    try:
        _sp = csv_path.stem.split("_"); _ys, _ye = _sp[-2], _sp[-1]
    except Exception:
        _ys, _ye = str(cfg.DATA_YEAR_START), str(cfg.DATA_YEAR_END)
    tip_master_path = Path(f"data_storage/members/TIP/ESG_MASTER_WIDE_TIP_MEMBERS_{_ys}_{_ye}.csv")
    _update_tip_members_file(csv_path, tip_master_path)
    _sync_company_member_files(master)
    _sync_consolidate_excel(master)

    # Parquet snapshot
    co_safe  = company.replace(" ", "_").replace("/", "_")
    ts       = datetime.now().strftime("%Y%m%d_%H%M%S")
    ver_dir  = Path("data_storage") / "versions" / co_safe
    ver_dir.mkdir(parents=True, exist_ok=True)
    filename = f"{co_safe}_{year}_elec_{ts}.parquet"
    single_row = master[(master["Company"] == company) & (master["Year"] == year)].copy()
    try:
        single_row.to_parquet(ver_dir / filename, index=False)
    except Exception:
        filename = "[parquet skipped]"

    return (f"Electricity saved — {len(updated_years)} year(s) updated "
            f"({min(updated_years)}-{max(updated_years)}) converted MWh to GJ. "
            f"Consolidate + member files synced. "
            f"Snapshot: versions/{co_safe}/{filename}")



def _sync_consolidate_excel(master_df: "pd.DataFrame") -> None:
    """
    Sync the CONSOLIDATED_DUMMY Excel (Raw Dummy data sheet) from the master wide CSV.

    The Raw Dummy data sheet stores data in long format: one row per
    (Company, Year, Row_Label). This function overwrites it completely from
    the current master wide DataFrame so that the consolidate stays in sync
    after any save from the platform.
    """
    from pathlib import Path
    from openpyxl import load_workbook

    xl_path = Path("data_storage/master/CONSOLIDATED_DUMMY_2009_2023.xlsx")
    if not xl_path.exists():
        return  # nothing to sync yet

    # Mapping: wide-CSV column  →  (Section, Row_Label)
    COL_MAP = {
        "Total no. of sites":                              ("ISO 14001",    "Total no. of sites"),
        "ISO 14001 sites":                                 ("ISO 14001",    "ISO 14001 sites"),
        "% certified sites":                               ("ISO 14001",    "% certified sites"),
        "Production":                                      ("Production",   "Production"),
        "Water intake":                                    ("Water",        "Water intake"),
        "Water intake - KPI":                              ("Water",        "Water intake - KPI"),
        "Total Electricity":                               ("Energy",       "Total Electricity"),
        "Renewable Electricity Purchased":                 ("Energy",       "Renewable Electricity Purchased"),
        "Non-Renewable Electricity Purchased":             ("Energy",       "Non-Renewable Electricity Purchased"),
        "Self-generated AND consumed electricity on-site": ("Energy",       "Self-generated AND consumed electricity on-site"),
        "Purchased Steam":                                 ("Energy",       "Purchased Steam"),
        "Sold Electricity":                                ("Energy",       "Sold Electricity"),
        "Sold Steam":                                      ("Energy",       "Sold Steam"),
        "Natural Gas":                                     ("Energy",       "Natural Gas"),
        "Coal":                                            ("Energy",       "Coal"),
        "Propane":                                         ("Energy",       "Propane"),
        "Fuel Oil":                                        ("Energy",       "Fuel Oil"),
        "Diesel":                                          ("Energy",       "Diesel"),
        "Petrol":                                          ("Energy",       "Petrol"),
        "Biomass":                                         ("Energy",       "Biomass"),
        "Waste tires":                                     ("Energy",       "Waste tires"),
        "LPG":                                             ("Energy",       "LPG"),
        "Other":                                           ("Energy",       "Other"),
        "Total energy":                                    ("Energy",       "Total energy"),
        "Total energy - KPI":                              ("Energy",       "Total energy - KPI"),
        "Total CO2 - Scope 1":                             ("CO2 emissions","Total CO2 - Scope 1"),
        "Total CO2 - Scope 2":                             ("CO2 emissions","Total CO2 - Scope 2"),
        "Total CO2":                                       ("CO2 emissions","Total CO2"),
        "Total CO2 - KPI":                                 ("CO2 emissions","Total CO2 - KPI"),
        "Total Waste":                                     ("Waste",        "Total Waste"),
        "Waste Recovered":                                 ("Waste",        "Waste Recovered"),
        "Recovery Rate":                                   ("Waste",        "Recovery Rate"),
        **{_elec_col(c): ("Energy", f"Electricity - {c}") for c in ELEC_ALL_COUNTRIES},
        # People & Governance (promoted from supplementary)
        "Stress Water Withdrawal":     ("Water",       "Stress water withdrawal"),
        "Non-Stress Water Withdrawal": ("Water",       "Non-stress water withdrawal"),
        "Coal Sub-Bituminous":         ("Energy",      "Coal — Sub-bituminous"),
        "Coal Brown Briquettes":       ("Energy",      "Coal — Brown briquettes"),
        "Coal Other Bituminous":       ("Energy",      "Coal — Other bituminous"),
        "HS External Audit Sites":     ("H&S",         "Externally audited H&S sites"),
        "HS Internal Audit Sites":     ("H&S",         "Internally audited H&S sites"),
        "HS External Audit %":         ("H&S",         "H&S external audit coverage %"),
        "HS Internal Audit %":         ("H&S",         "H&S internal audit coverage %"),
        "Total Employees":             ("Diversity",   "Total employees"),
        "Female Employees":            ("Diversity",   "Female employees"),
        "Female Employees %":          ("Diversity",   "% Female employees"),
        "Board Total":                 ("Diversity",   "Board of Directors total"),
        "Female Board":                ("Diversity",   "Female Board members"),
        "Female Board %":              ("Diversity",   "% Female Board"),
        "SBT Total":                   ("SBT",         "Total with science-based target"),
        "SBT Validated":               ("SBT",         "SBT — Validated"),
        "SBT Committed":               ("SBT",         "SBT — Committed"),
        "SBT Non-Committed":           ("SBT",         "SBT — Non-committed"),
    }

    # Build long rows from master_df
    long_rows = []  # list of dicts: Company, Row, Year, Data, Section, Row_Label, Notes, Consistency test
    row_order = list(COL_MAP.keys())
    # Assign fixed row numbers to match what build_esg_master.py uses
    ROW_NUM = {col: i + 1 for i, col in enumerate(row_order)}

    # Pre-compute which electricity country columns have any non-zero value
    # across the whole master — only those countries get rows in the consolidate.
    active_elec_cols = {
        col for col in COL_MAP
        if col.startswith("Elec_") and col.endswith("_GJ")
        and col in master_df.columns
        and master_df[col].fillna(0).ne(0).any()
    }

    for _, wrow in master_df.sort_values(["Company", "Year"]).iterrows():
        company = wrow["Company"]
        year    = int(wrow["Year"]) if pd.notna(wrow.get("Year")) else None
        if not company or not year:
            continue
        for col, (section, label) in COL_MAP.items():
            # Skip electricity country columns that are all-zero across the dataset
            is_elec_country = col.startswith("Elec_") and col.endswith("_GJ")
            if is_elec_country and col not in active_elec_cols:
                continue
            val = wrow.get(col)
            # For an active electricity country, skip rows where this company-year is zero
            if is_elec_country and (pd.isna(val) or float(val) == 0):
                continue
            long_rows.append({
                "Company": company,
                "Row":     ROW_NUM[col],
                "Year":    year,
                "Data":    float(val) if pd.notna(val) else None,
                "Section": section,
                "Row_Label": label,
                "Notes":   None,
                "Consistency test": None,
            })

    if not long_rows:
        return

    try:
        wb = load_workbook(xl_path)
        ws = wb["Raw Dummy data"]
        # Clear existing data rows (keep header row 1)
        for r in range(2, ws.max_row + 1):
            for c in range(1, 9):
                ws.cell(r, c).value = None
        # Write new rows
        cols = ["Company", "Row", "Year", "Data", "Section", "Row_Label", "Notes", "Consistency test"]
        for i, row in enumerate(long_rows):
            for j, col in enumerate(cols):
                ws.cell(i + 2, j + 1).value = row[col]
        wb.save(xl_path)
    except (PermissionError, OSError):
        pass  # file locked — skip, master CSV is the source of truth
    except Exception:
        pass  # any other error is also non-fatal


def _elec_col(country: str) -> str:
    """Canonical master CSV column name for a country's electricity (GJ)."""
    return "Elec_" + country.replace(" ", "_") + "_GJ"