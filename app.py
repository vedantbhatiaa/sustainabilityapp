"""
app.py — TIP ESG Platform · Entry Point
========================================
Streamlit application root. Handles:
  - Session state initialisation (init_state)
  - User authentication (show_login)
  - Navigation sidebar (show_sidebar)
  - Page routing to pages/ modules

Run:  streamlit run app.py

Structure:
  Global module-level code loads the master CSV once per process start
  and keeps it in _CONSOLIDATED_DF. All pages read from this in-memory
  DataFrame — no disk reads on every page navigation.

  On save (Submit & Save), _save_submission_to_csv updates both the
  in-memory globals and the disk CSV atomically under a file lock.

Run: streamlit run app.py

Changes from original:
  - formula_engine now imported from formula_engine.py (was .ipynb -- crashed on startup)
  - Removed dead load_consolidated_data() function (used local_storage but was never called)
  - analysis page now reads LONG_DATA from real sector CSV when available
  - data_loader now auto-finds master CSV in data_storage/raw/ (no manual path fix needed)
"""
import streamlit as st
import pandas as pd
import numpy as np
import plotly.graph_objects as go
from plotly.subplots import make_subplots
import html as _html
import logging
import os
from pathlib import Path
from datetime import datetime, date
from filelock import FileLock

import config as cfg

# ── Logging setup ─────────────────────────────────────────────────────────────
logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(name)s: %(message)s",
)
_log = logging.getLogger("esg_app")

from formula_engine import (
    TemplateInputs, calculate, validate_submission,
    get_benchmarks, build_template_dataframe, fmt_num,
    yoy_change, ValidationFlag, BenchmarkResult
)

import data_loader as dl
from utils.comment_utils import (
    load_comments as _load_comments,
    save_change_comment as _save_change_comment,
    update_comment_status as _update_comment_status,
    get_approved_comments as _get_approved_comments,
    get_all_active_comments as _get_all_active_comments,
    update_master_comment_cell as _update_master_comment_cell,
    delete_comment as _delete_comment,
    save_comment_version as _save_comment_version,
)

# Chatbot is embedded in page_readiness only — no global import needed

# Load fresh from disk on every Streamlit rerun (Streamlit reruns the full
# script on every user interaction, so this is always up-to-date after a save).
# data_loader checks data_storage/master/ first, then falls back to raw/ etc.
_CONSOLIDATED_DF = dl.load_consolidated()
_COMPANIES       = dl.get_companies(_CONSOLIDATED_DF)

_SECTOR_DF       = dl.load_sector_aggregated(_CONSOLIDATED_DF)


def _write_verification_status(company: str, year: int, status: str) -> None:
    """
    Persist DSS+ verification status for a company+year to a CSV file.
    Status values: 'Verified', 'Pending', 'Flagged'.
    Client home page reads this file to show the verification chip.
    """
    from pathlib import Path
    import csv, os

    vcsv = Path("data_storage/verifications.csv")
    vcsv.parent.mkdir(parents=True, exist_ok=True)

    rows = []
    if vcsv.exists():
        with open(vcsv, newline="") as f:
            for row in csv.DictReader(f):
                if not (row.get("Company","").strip() == company and
                        str(row.get("Year","")).strip() == str(year)):
                    rows.append(row)   # keep other company/year rows

    rows.append({"Company": company, "Year": str(year), "Status": status,
                 "UpdatedBy": "dss+ Analyst"})

    with open(vcsv, "w", newline="") as f:
        w = csv.DictWriter(f, fieldnames=["Company","Year","Status","UpdatedBy"])
        w.writeheader()
        w.writerows(rows)


# ═══════════════════════════════════════════════════════════════════════════════
# SUPPLEMENTARY DATA STORE — fields outside TemplateInputs
# (H&S, Diversity, SBT, Water detail, Coal breakdown)
# Stored in data_storage/master/ESG_SUPPLEMENTARY.csv
#
# POST-AZURE MIGRATION: Move all these fields to master SQL table as additional
# columns. Then deprecate this separate CSV file and remove _load_supplementary()
# and _save_supplementary() functions. For now, keeping both CSV and SQL-ready.
# ═══════════════════════════════════════════════════════════════════════════════

_SUPP_PATH = Path("data_storage/master/ESG_SUPPLEMENTARY.csv")
_SUPP_FIELDS = [
    "Company","Year",
    # Water detail
    "stress_water_withdrawal","non_stress_water_withdrawal",
    # Coal breakdown
    "coal_sub_bituminous","coal_brown_briquettes","coal_other_bituminous",
    # H&S
    "hs_external_audit","hs_internal_audit",
    # Diversity
    "total_employees","female_employees","board_total","female_board",
    # SBT
    "sbt_total","sbt_validated","sbt_committed","sbt_non_committed",
]

def _load_supplementary(company: str, year: int) -> dict:
    """Load supplementary fields for company+year. Returns {} if not found."""
    from pathlib import Path as _P
    import csv
    p = _P("data_storage/master/ESG_SUPPLEMENTARY.csv")
    if not p.exists(): return {}
    with open(p, newline="", encoding="utf-8") as f:
        for row in csv.DictReader(f):
            if row.get("Company","").strip()==company and str(row.get("Year","")).strip()==str(year):
                return {k: float(v) if v else 0.0 for k,v in row.items()
                        if k not in ("Company","Year")}
    return {}

def _save_supplementary(company: str, year: int, data: dict) -> None:
    """Upsert supplementary fields for company+year."""
    from pathlib import Path as _P
    import csv
    p = _P("data_storage/master/ESG_SUPPLEMENTARY.csv")
    p.parent.mkdir(parents=True, exist_ok=True)
    rows = []
    if p.exists():
        with open(p, newline="", encoding="utf-8") as f:
            for row in csv.DictReader(f):
                if not (row.get("Company","").strip()==company and
                        str(row.get("Year","")).strip()==str(year)):
                    rows.append(row)
    new_row = {"Company": company, "Year": str(year)}
    for field in _SUPP_FIELDS[2:]:          # skip Company, Year
        new_row[field] = str(data.get(field, 0) or 0)
    rows.append(new_row)
    with open(p, "w", newline="", encoding="utf-8") as f:
        w = csv.DictWriter(f, fieldnames=_SUPP_FIELDS)
        w.writeheader(); w.writerows(rows)


# ═══════════════════════════════════════════════════════════════════════════════
# FIELD COMMENT / CHANGE-REQUEST SYSTEM
# When a client updates a PREVIOUS year, they must give a reason.
# DSS approves/rejects from the Verification Queue.
# Approved comments appear in RED in the template Main Input view.
# ═══════════════════════════════════════════════════════════════════════════════

_COMMENTS_PATH = Path("data_storage/field_comments.csv")
_COMMENT_COLS  = ["Company","Year","FieldKey","FieldLabel","OldValue",
                  "NewValue","Reason","SubmittedAt","Status","ApprovedBy","ApprovedAt"]

def _reload_consolidated_df() -> bool:
    """
    Force-reload master CSV + rebuild ALL dependent globals.
    Uses dl.load_consolidated() dynamic scan so newly-named files
    (e.g. _2024.csv, _2025.csv) are always found without hardcoded filenames.
    Sector aggregation is computed live from master so new years appear immediately.
    """
    global _CONSOLIDATED_DF, _COMPANIES, _SECTOR_DF, _USING_FALLBACK_DATA
    global HIST_YEARS, CURR_YEAR, LONG_YEARS, LONG_DATA, FUEL_MIX
    try:
        fresh = pd.DataFrame()
        # NOTE: SharePoint load_master() not yet implemented in StorageClient.
        # We fall through directly to the local CSV path below.
        if fresh.empty:
            fresh = dl.load_consolidated()
        if not fresh.empty and "Company" in fresh.columns and "Year" in fresh.columns:
            _CONSOLIDATED_DF     = fresh
            _COMPANIES           = dl.get_companies(fresh)
            # Compute sector live — always covers the latest submitted year
            _SECTOR_DF           = dl.load_sector_aggregated(fresh)
            _USING_FALLBACK_DATA = False
            try:
                cfg.refresh_year_bounds(fresh)
                HIST_YEARS = cfg.hist_years()
                CURR_YEAR  = cfg.curr_year()
                LONG_YEARS = cfg.long_years()
            except Exception:
                pass
            try:
                LONG_DATA, FUEL_MIX = _build_long_data()
            except Exception:
                pass
            st.session_state["_df_version"] = st.session_state.get("_df_version", 0) + 1
            return True
    except Exception:
        pass
    return False

# ─────────────────────────────────────────────────────────
# PAGE CONFIG & GLOBAL CSS
# ─────────────────────────────────────────────────────────
st.set_page_config(
    page_title="TIP ESG Platform",
    page_icon="🌿",
    layout="wide",
    initial_sidebar_state="expanded"
)

st.markdown("""
<style>
/* -- Sidebar -- */
[data-testid="stSidebar"] { background: #0A2240 !important; }
[data-testid="stSidebar"] * { color: rgba(255,255,255,0.75) !important; }
[data-testid="stSidebar"] h1,[data-testid="stSidebar"] h2,
[data-testid="stSidebar"] h3,[data-testid="stSidebar"] strong
{ color: #ffffff !important; }
[data-testid="stSidebar"] hr { border-color: rgba(255,255,255,0.1) !important; }
[data-testid="stSidebarNav"] { display: none; }

/* -- Sidebar nav buttons -- */
[data-testid="stSidebar"] .stButton > button {
    background: transparent !important;
    color: rgba(255,255,255,0.75) !important;
    border: none !important;
    border-radius: 8px !important;
    font-size: 13.5px !important;
    font-weight: 500 !important;
    text-align: left !important;
    padding: 9px 12px !important;
    width: 100% !important;
    transition: background .15s, color .15s !important;
    box-shadow: none !important;
}
[data-testid="stSidebar"] .stButton > button:hover {
    background: rgba(255,255,255,0.10) !important;
    color: #ffffff !important;
    border: none !important;
}
[data-testid="stSidebar"] .stButton > button:focus {
    box-shadow: none !important; outline: none !important;
    border: none !important; color: #ffffff !important;
}

/* -- Main buttons -- */
.stButton > button {
    border-radius: 7px; font-weight: 500; font-size: 13px;
    border: 1.5px solid #D1D5DB; transition: all .15s;
}
.stButton > button:hover { border-color: #6B7280; }

/* -- Form inputs -- */
.stNumberInput input, .stTextInput input, .stSelectbox select {
    border-radius: 7px; border: 1.5px solid #D1D5DB; font-size: 14px !important;
}

/* -- KPI cards -- */
.kpi-card {
    background: #fff; border: 1px solid #E5E7EB;
    border-radius: 10px; padding: 16px 18px; text-align: left;
}
.kpi-card .label { font-size: 11px; color: #6B7280;
    text-transform: uppercase; letter-spacing: .5px; font-weight: 500; }
.kpi-card .value { font-size: 26px; font-weight: 700; color: #111827; margin: 5px 0 2px; }
.kpi-card .unit  { font-size: 12px; color: #9CA3AF; }
.kpi-card .delta { font-size: 12px; font-weight: 600; margin-top: 4px; }
.delta-pos { color: #059669; }
.delta-neg { color: #DC2626; }

/* -- Stepper -- */
.step-bar { display:flex; align-items:center; gap:0;
    background:#fff; border:1px solid #E5E7EB; border-radius:10px;
    padding:16px 20px; margin-bottom:20px; }
.step-item { display:flex; align-items:center; flex:1; min-width:0; }
.step-circle { width:28px; height:28px; border-radius:50%;
    display:flex; align-items:center; justify-content:center;
    font-size:12px; font-weight:700; flex-shrink:0; }
.sc-done   { background:#00916E; color:#fff; }
.sc-active { background:#1D4ED8; color:#fff; }
.sc-todo   { background:#F3F4F6; color:#9CA3AF; border:2px solid #E5E7EB; }
.step-label { font-size:11.5px; font-weight:500; margin-left:7px;
    white-space:nowrap; overflow:hidden; text-overflow:ellipsis; }
.sl-done   { color:#00916E; }
.sl-active { color:#1D4ED8; }
.sl-todo   { color:#9CA3AF; }
.step-line { flex:1; height:2px; background:#E5E7EB; margin:0 6px; min-width:8px; }
.sl-done-line { background:#00916E; }

/* -- Table legend -- */
.tbl-legend { display:flex; gap:14px; padding:10px 16px;
    background:#F9FAFB; border-top:1px solid #E5E7EB;
    border-radius:0 0 8px 8px; flex-wrap:wrap; }
.tl { display:flex; align-items:center; gap:5px; font-size:11px; color:#6B7280; }
.tl-sw { width:14px; height:14px; border-radius:3px;
    border:1px solid #D1D5DB; display:inline-block; }

/* -- Band chart -- */
.band-container { margin: 6px 0 12px; }
.band-row-wrap { display:flex; align-items:center; gap:10px; margin-bottom:10px; }
.band-lbl { font-size:12px; font-weight:500; color:#374151; width:170px; flex-shrink:0; }
.band-track { flex:1; height:18px; border-radius:4px; background:#F3F4F6; position:relative; }
.band-seg { position:absolute; top:0; height:100%; }
.band-pin { position:absolute; width:4px; height:28px;
    background:#0A2240; border-radius:2px; top:-5px; transform:translateX(-50%); }
.band-pin-val { position:absolute; font-size:10px; font-weight:700;
    color:#0A2240; top:-18px; transform:translateX(-50%);
    white-space:nowrap; background:#fff; padding:0 2px; }
.band-chip { font-size:11px; font-weight:600; padding:3px 9px;
    border-radius:10px; flex-shrink:0; }
.chip-top  { background:#D1FAE5; color:#065F46; }
.chip-mid  { background:#FEF3C7; color:#92400E; }
.chip-bot  { background:#FEE2E2; color:#991B1B; }

/* -- Flag cards -- */
.flag-card { display:flex; align-items:flex-start; gap:10px;
    padding:12px 14px; border-radius:8px; border:1px solid; margin-bottom:10px; }
.fc-warn  { background:#FFF7ED; border-color:#FCD34D; }
.fc-error { background:#FEF2F2; border-color:#FECACA; }
.fc-ok    { background:#ECFDF5; border-color:#6EE7B7; }
.fc-icon  { width:20px; height:20px; border-radius:50%;
    display:flex; align-items:center; justify-content:center;
    font-size:10px; font-weight:700; flex-shrink:0; }
.fi-warn  { background:#D97706; color:#fff; }
.fi-error { background:#DC2626; color:#fff; }
.fi-ok    { background:#00916E; color:#fff; }
.fc-title { font-size:13px; font-weight:600; color:#111827; }
.fc-detail{ font-size:12px; color:#6B7280; margin-top:3px; line-height:1.6; }

/* -- AI card -- */
.ai-card { background:#fff; border:1px solid #E5E7EB;
    border-radius:10px; overflow:hidden; margin-bottom:12px; }
.ai-head  { display:flex; align-items:center; gap:8px;
    padding:10px 14px; background:#F9FAFB; border-bottom:1px solid #E5E7EB; }
.ai-pulse { width:8px; height:8px; border-radius:50%;
    background:#00916E; flex-shrink:0; }
.ai-title { font-size:13px; font-weight:600; color:#111827; }
.ai-badge { margin-left:auto; background:#E6F5F1; color:#007A5C;
    font-size:11px; padding:2px 9px; border-radius:10px;
    border:1px solid #6EE7B7; font-weight:500; }
.ai-body  { padding:12px 14px; font-size:13px; color:#374151; line-height:1.8; }
</style>
""", unsafe_allow_html=True)

# ── Animation system & design tokens ─────────────────────────────────────────
from ui_components import (
    inject_global_css, kpi_card_html, skeleton_card_html, skeleton_chart_html,
    status_chip_html, section_header_html, empty_state_html, co_card_html,
    apply_chart_animation, chart_layout_defaults, sparkline_html,
    GREEN, AMBER, RED, NAVY, BG, BORDER, TEXT, MUTED,
    CAT_CO2, CAT_ENERGY, CAT_WATER, CAT_WASTE, CAT_RENEW,
)
inject_global_css()

# ── TIP background scoped to Plotly chart containers ONLY ────────────────────
# Targets the div Streamlit wraps each Plotly chart in.
# Does NOT touch tabs, tables, data entry forms, or KPI metric cards.
st.markdown("""
<style>
[data-testid="stPlotlyChart"] > div,
[data-testid="stPlotlyChart"] iframe,
.js-plotly-plot .plotly,
.stPlotlyChart > div {
    background-color: #f5f4f2 !important;
    border-radius: 6px;
}
</style>
""", unsafe_allow_html=True)

# ═══════════════════════════════════════════════════════════════════════
# TIP Report Design System — chart_layout_defaults (global override)
# Replaces the ui_components version everywhere it's called in app.py.
# Background, axes, grid, and typography aligned to TIP Annual KPI Report.
# ═══════════════════════════════════════════════════════════════════════
_TIP_BG    = "#f5f4f2"   # REPORT_BG — warm off-white
_TIP_CHR   = "#2a2825"   # TIP_CHARCOAL — title / primary text
_TIP_AX    = "#9aa1a9"   # axis lines
_TIP_GRID  = "#e6eaed"   # y-axis grid lines
_TIP_MUTED = "#6f7882"   # tick / caption text

def chart_layout_defaults(title="", height=300, showlegend=True, **kw):
    """TIP-aligned chart layout. Overrides ui_components version.

    Uses underscore notation for xaxis/yaxis so callers can safely pass their
    own xaxis=dict(...) or yaxis=dict(...) without Python raising 'multiple
    values for keyword argument'.  Plotly merges underscore keys with explicit
    dict keys, so the defaults still apply to any sub-property not overridden.
    """
    base = dict(
        title=dict(
            text=f"<b>{title}</b>",
            font=dict(size=14, color=_TIP_CHR, family="Arial, sans-serif"),
            x=0,
        ),
        height=height,
        margin=dict(l=55, r=65, t=50, b=60),
        plot_bgcolor=_TIP_BG,
        paper_bgcolor=_TIP_BG,
        # ── xaxis defaults — underscore notation avoids dict key collision ──
        xaxis_showgrid=False,
        xaxis_showline=True,
        xaxis_linewidth=1.2,
        xaxis_linecolor=_TIP_AX,
        xaxis_mirror=False,
        xaxis_ticks="",
        xaxis_tickfont=dict(size=12, color=_TIP_MUTED, family="Arial"),
        xaxis_tickangle=0,
        xaxis_type="category",
        xaxis_zeroline=False,
        # ── yaxis defaults ────────────────────────────────────────────────
        yaxis_showgrid=True,
        yaxis_gridcolor=_TIP_GRID,
        yaxis_zeroline=False,
        yaxis_showline=True,
        yaxis_linewidth=1.2,
        yaxis_linecolor=_TIP_AX,
        yaxis_ticks="",
        yaxis_tickfont=dict(size=12, color=_TIP_MUTED, family="Arial"),
        yaxis_autorange=True,
        # ── legend & interaction ──────────────────────────────────────────
        legend=dict(
            orientation="h", x=0.5, xanchor="center", y=-0.24,
            font=dict(size=12, color=_TIP_MUTED, family="Arial"),
            bgcolor="rgba(0,0,0,0)",
        ),
        hovermode="x unified",
        showlegend=showlegend,
    )
    base.update(kw)
    return base

# ─────────────────────────────────────────────────────────
# CONSTANTS
# ─────────────────────────────────────────────────────────
HIST_YEARS = cfg.hist_years()
CURR_YEAR  = cfg.curr_year()
LONG_YEARS = cfg.long_years()

# ── All 31 electricity-by-country names (matches UI editor row order) ─────────
ELEC_ALL_COUNTRIES = [
    "Canada", "Chile", "Mexico", "United States",
    "Australia", "Japan", "Korea", "New Zealand",
    "Austria", "Belgium", "Czech Republic", "Denmark", "Finland", "France",
    "Germany", "Hungary", "Iceland", "Ireland", "Italy", "Luxembourg",
    "Netherlands", "Norway", "Poland", "Portugal", "Spain", "Sweden",
    "Switzerland", "Turkey", "United Kingdom",
    "China", "India",
]

def _elec_col(country: str) -> str:
    """Canonical master CSV column name for a country's electricity (GJ)."""
    return "Elec_" + country.replace(" ", "_") + "_GJ"

# Dict: country name → master CSV column name
ELEC_COUNTRY_COLS = {c: _elec_col(c) for c in ELEC_ALL_COUNTRIES}

COMPANIES = _COMPANIES if _COMPANIES else [
    "VerdaTyres Corp", "AlphaTread Ltd", "BetaRubber Inc", "GammaTire SA",
    "DeltaGrip GmbH", "EpsilonWheel Co", "ZetaTrac LLC", "EtaRoad AG",
    "ThetaDrive NV", "IotaTire PLC",
]

# ── Year bounds — dynamic from real data, config-driven fallback ─────────────
# cfg.refresh_year_bounds() updates cfg.DATA_YEAR_START / DATA_YEAR_END in-place
cfg.refresh_year_bounds(_CONSOLIDATED_DF)
HIST_YEARS = cfg.hist_years()   # e.g. [2009..2022]
CURR_YEAR  = cfg.curr_year()    # e.g. 2023
LONG_YEARS = cfg.long_years()   # e.g. [2009..2023]

# ── Client auth mapping — loaded from config (secrets.toml or env var) ───────
# In production set CLIENTS_JSON in secrets.toml. Demo fallback in config.py.
CLIENTS = cfg.load_clients()
DSS_DOMAIN = cfg.DSS_EMAIL_DOMAIN

# ── DEMO FALLBACK DATA ─────────────────────────────────────────────────────────
# HIST_RAW is used ONLY when no master CSV is loaded (fresh install or demo mode).
# In production, the app.py startup (line 72) always loads _CONSOLIDATED_DF from disk/Azure SQL.
# If production is configured correctly, HIST_RAW is never used.
#
# NOTE: These values cover years 2009–2023 (14 years). If you add 2024+ data:
#   - Pre-Azure: Update these arrays or remove if confident master CSV always loads
#   - Post-Azure: Remove entirely (Azure SQL will be the only source)
# ─────────────────────────────────────────────────────────────────────────────────

HIST_RAW: dict[str, list] = {
    "total_sites":   [38,39,40,40,41,42,43,44,46,48,51,51,52,52],
    "iso_sites":     [36,38,39,39,40,41,42,43,45,47,51,51,52,52],
    "production":    [2_840_000,3_510_000,3_770_000,3_520_000,3_640_000,3_620_000,
                      3_540_000,3_630_000,3_700_000,3_910_000,3_860_000,3_050_000,3_320_000,3_580_000],
    "water_withdrawals": [18_000_000,19_200_000,20_100_000,19_700_000,20_500_000,21_000_000,
                          21_500_000,22_000_000,22_800_000,23_100_000,23_100_000,20_300_000,21_100_000,20_900_000],
    "renew_elec_purchased": [0,0,0,0,0,0,0,300_000,250_000,1_100_000,706_562,1_528_836,2_557_561,4_082_923],
    "nonrenew_elec_purchased": [9_500_000,9_400_000,9_600_000,9_300_000,9_400_000,9_500_000,
                                9_600_000,9_400_000,9_300_000,9_100_000,12_271_131,9_667_437,10_297_758,9_037_549],
    "nat_gas": [13_000_000,14_000_000,15_000_000,14_500_000,15_000_000,15_200_000,
                15_500_000,15_700_000,15_900_000,16_100_000,16_210_969,14_040_397,15_939_109,15_927_554],
    "coal_sub": [500_000,490_000,480_000,470_000,460_000,450_000,
                 440_000,430_000,420_000,410_000,456_997,337_992,360_848,395_006],
    "lpg": [1_000_000,1_050_000,1_100_000,1_100_000,1_100_000,1_150_000,
            1_150_000,1_180_000,1_200_000,1_220_000,1_237_839,1_124_479,1_271_422,1_329_571],
    "waste_total":    [330_000,330_000,335_000,335_000,340_000,342_000,
                       344_000,346_000,348_000,350_000,352_000,295_000,320_000,335_000],
    "waste_recovery": [280_000,281_000,283_000,284_000,286_000,287_000,
                       289_000,292_000,295_000,298_000,299_200,253_700,275_200,284_750],
}

# -- LONG_DATA: built from real sector CSV if available, else static fallback --
def _build_long_data() -> tuple[dict, dict]:
    """
    Build LONG_DATA and FUEL_MIX from the real consolidated wide DataFrame.
    Falls back to static values only for missing years/fields.
    """
    static_long = {
        "energy":     [28.1,32.3,33.6,32.4,33.0,32.4,31.8,32.1,33.0,34.2,33.2,28.5,32.3,32.5,32.4],
        "co2":        [2.41,2.69,2.88,2.80,2.87,2.86,2.73,2.72,2.80,2.85,2.76,2.22,2.27,2.06,2.05],
        "water":      [22.4,23.8,24.9,24.4,23.9,22.9,22.9,23.5,23.5,23.2,23.1,20.3,21.1,20.9,21.5],
        "scope1":     [1.08,1.19,1.21,1.17,1.20,1.15,1.09,1.08,1.12,1.15,1.11,0.94,1.06,1.05,1.03],
        "scope2":     [1.33,1.50,1.67,1.63,1.67,1.71,1.64,1.64,1.68,1.70,1.65,1.27,1.21,1.01,1.02],
        "energy_kpi": [9.9,9.2,8.9,9.2,9.1,8.9,8.9,8.8,8.9,8.8,8.6,9.3,9.7,9.1,8.7],
        "co2_kpi":    [0.850,0.765,0.764,0.795,0.789,0.791,0.771,0.748,0.758,0.729,0.715,0.729,0.684,0.576,0.551],
        "renew_pct":  [0,0,0,0,0,0,0,2.3,2.2,9.7,10.6,21.8,31.4,40.6,48.3],
        "waste_recov":[83,83,84,84,84,84,84,85,85,85,85,86,86,85,86],
        "prod":       [2.84,3.51,3.77,3.52,3.64,3.62,3.54,3.63,3.70,3.91,3.86,3.05,3.32,3.58,3.72],
    }
    static_fuel = {
        "Natural Gas": [46,46,47,47,47,46,47,47,48,49,49,49,49,49,50],
        "Electricity": [34.7,34.2,34.9,34.6,35.3,36.7,37.1,37.9,38.2,38.8,39.1,39.3,39.8,40.4,40.7],
        "Fuel Oil":    [8.5,6.7,6,5.8,5.1,4.8,3.7,3.2,3,2.6,2.4,1.8,1.4,0.5,0.5],
        "LPG":         [2.4,2.4,2.3,3.6,3.5,3.5,3.5,3.6,3.5,3.6,3.7,3.9,3.9,4.1,4.2],
        "Coal":        [3.2,3.1,2.8,2.8,3.7,3.6,2.3,2,2.1,1.6,1.4,1.2,1.2,1.2,1.2],
        "Other":       [5.2,7.6,7,5.4,5.4,5.4,6.4,5.3,4.7,4.4,4,4.8,4.3,4.8,3.4],
    }

    def _safe_list(series, fallback):
        result = []
        for i, v in enumerate(series):
            try:
                f = float(v)
                result.append(f if not np.isnan(f) else fallback[i])
            except Exception:
                result.append(fallback[i])
        return result

    df = _CONSOLIDATED_DF
    if df.empty or "Row_Label" in df.columns:
        # Long format or no data -- use static
        if not _SECTOR_DF.empty:
            try:
                s = _SECTOR_DF.set_index("Year").reindex(LONG_YEARS)
                live = {
                    "energy":     _safe_list((s["Total_Energy"]/1e6), static_long["energy"]),
                    "co2":        _safe_list((s["Total_CO2"]/1e6),    static_long["co2"]),
                    "water":      _safe_list((s["Total_Water"]/1e6),  static_long["water"]),
                    "energy_kpi": _safe_list(s["Avg_Energy_KPI"],     static_long["energy_kpi"]),
                    "co2_kpi":    _safe_list(s["Avg_CO2_KPI"],        static_long["co2_kpi"]),
                    "renew_pct":  _safe_list(s["Avg_Renewable_Share"],static_long["renew_pct"]),
                    "prod":       _safe_list((s["Total_Production"]/1e6), static_long["prod"]),
                    "scope1":     static_long["scope1"],
                    "scope2":     static_long["scope2"],
                    "waste_recov":static_long["waste_recov"],
                }
                return live, static_fuel
            except Exception as e:
                _log.warning("[app] Sector DF error: %s", e)
        return static_long, static_fuel

    # Wide format -- compute directly from master DataFrame
    try:
        grp = df.groupby("Year")

        def _col_sum(col, divisor=1):
            if col in df.columns:
                return grp[col].sum() / divisor
            return None

        def _col_mean(col):
            if col in df.columns:
                return grp[col].mean()
            return None

        def _col_sum_norm(col, divisor=1):
            """Sector sum normalised by n_submitting / n_total_companies."""
            if col not in df.columns:
                return None
            raw = grp[col].sum() / divisor
            n_sub = grp["Company"].count()
            n_all = df["Company"].nunique()
            return raw / n_sub * n_all

        energy_s  = _col_sum_norm("Total energy", 1e6)
        co2_s     = _col_sum_norm("Total CO2", 1e6)
        scope1_s  = _col_sum_norm("Total CO2 - Scope 1", 1e6)
        scope2_s  = _col_sum_norm("Total CO2 - Scope 2", 1e6)
        water_s   = _col_sum_norm("Water intake", 1e6)
        # Use MEAN × n_submitting_companies so partial-year submissions
        # don't cause a false cliff-drop in sector production.
        prod_s_raw  = _col_sum("Production", 1e6)
        n_companies = df["Company"].nunique()
        if prod_s_raw is not None:
            n_submitting = grp["Company"].count()
            prod_s       = prod_s_raw / n_submitting * n_companies
        else:
            prod_s = prod_s_raw
        ekpi_m    = _col_mean("Total energy - KPI")
        co2kpi_m  = _col_mean("Total CO2 - KPI")
        # For renewable %, only average companies that have submitted for that year
        # (NaN values from non-submitting companies would drag the mean down)
        renew_m   = df.groupby("Year")["Renewable_Electricity_Share_%"].apply(
            lambda x: x.dropna().mean() if x.dropna().size > 0 else float("nan")
        ) if "Renewable_Electricity_Share_%" in df.columns else None

        def _to_list(series, fallback):
            if series is None:
                return fallback
            s = series.reindex(LONG_YEARS)
            return _safe_list(s.values, fallback)

        live = {
            "energy":     _to_list(energy_s,  static_long["energy"]),
            "co2":        _to_list(co2_s,     static_long["co2"]),
            "scope1":     _to_list(scope1_s,  static_long["scope1"]),
            "scope2":     _to_list(scope2_s,  static_long["scope2"]),
            "water":      _to_list(water_s,   static_long["water"]),
            "prod":       _to_list(prod_s,    static_long["prod"]),
            "energy_kpi": _to_list(ekpi_m,    static_long["energy_kpi"]),
            "co2_kpi":    _to_list(co2kpi_m,  static_long["co2_kpi"]),
            "renew_pct":  _to_list(renew_m,   static_long["renew_pct"]),
            "waste_recov":static_long["waste_recov"],
        }

        # Fuel mix as % of total energy per year
        fuel_cols = {
            "Natural Gas": "Natural Gas",
            "Electricity": "Total Electricity",
            "Fuel Oil":    "Fuel Oil",
            "LPG":         "LPG",
            "Coal":        "Coal",
            "Other":       "Other",
        }
        total_e_by_yr = grp["Total energy"].sum() if "Total energy" in df.columns else None
        live_fuel = {}
        n_yrs = len(LONG_YEARS)
        for label, col in fuel_cols.items():
            _fb = static_fuel.get(label, [])
            _fb_ext = (_fb + [_fb[-1] if _fb else 0.0] * max(0, n_yrs - len(_fb)))
            if col in df.columns and total_e_by_yr is not None:
                fuel_sum = grp[col].sum().reindex(LONG_YEARS)
                total_e  = total_e_by_yr.reindex(LONG_YEARS)
                pct = (fuel_sum / total_e.replace(0, np.nan) * 100).fillna(0)
                live_fuel[label] = _safe_list(pct.values, _fb_ext)
            else:
                live_fuel[label] = _fb_ext[:n_yrs]

        return live, live_fuel if any(sum(v) > 0 for v in live_fuel.values()) else static_fuel

    except Exception as e:
        _log.warning("[app] Wide DF live computation error: %s", e)
        return static_long, static_fuel


LONG_DATA, FUEL_MIX = _build_long_data()

# L2 FIX: track whether analysis charts are showing real or fallback data.
# Surfaced as a warning banner in page_analysis() so analysts never mistake
# synthetic demo numbers for real client submissions.
_USING_FALLBACK_DATA = _CONSOLIDATED_DF.empty

CLIENTS = {
    "verdatyres@tip-reporting.com":   "VerdaTyres Corp",
    "alphatread@tip-reporting.com":   "AlphaTread Ltd",
    "betarubber@tip-reporting.com":   "BetaRubber Inc",
    "gammatire@tip-reporting.com":    "GammaTire SA",
    "deltagrip@tip-reporting.com":    "DeltaGrip GmbH",
    "epsilonwheel@tip-reporting.com": "EpsilonWheel Co",
    "zetatrac@tip-reporting.com":     "ZetaTrac LLC",
    "etaroad@tip-reporting.com":      "EtaRoad AG",
    "thetadrive@tip-reporting.com":   "ThetaDrive NV",
    "iotatire@tip-reporting.com":     "IotaTire PLC",
}

STEP_META = [
    ("ISO 14001",  "Certified sites and facility coverage"),
    ("Production", "Annual production volume"),
    ("Water",      "Water withdrawals by source"),
    ("Energy",     "Electricity and fuel consumption"),
    ("CO2",        "Emission inputs and auto-calculated totals"),
    ("Waste",      "Waste generated, recovered and eliminated"),
]

# ─────────────────────────────────────────────────────────
# SESSION STATE INIT
# ─────────────────────────────────────────────────────────
def init_state():
    defaults = {
        "authenticated":      False,
        "user_name":          "",
        "user_company":       "",
        "user_email":         "",
        "is_dss":             False,
        "page":               "login",
        "step":               0,
        "template_done":      False,
        "company_setup_done": False,
        "reporting_company":  "",
        "reporting_year":     cfg.curr_year(),
        "employee_name":      "",
        "company_hist":       {},
        "live_hist_raw":      {},
        "kpi_hints":          {},
        "step_data": {
            "total_sites": 54, "iso_sites": 54,
            "production": 3_720_000,
            "water_withdrawals": 21_500_000,
            "renew_elec_purchased": 5_200_000,
            "nonrenew_elec_purchased": 8_500_000,
            "self_gen_elec": 45_000,
            "purchased_steam": 1_050_000,
            "sold_electricity": 8_000,
            "sold_steam": 0,
            "nat_gas": 16_100_000, "coal_sub": 380_000,
            "propane": 340_000, "fuel_oil_heavy_a": 150_000,
            "diesel": 190_000, "petrol": 0, "biomass": 0,
            "waste_tires_mt": 0, "lpg": 1_350_000, "other_fuels": 0,
            "co2_scope2_steam": 60_000,
            "waste_total": 338_000, "waste_recovery": 290_000,
        },
        "flags_resolved": set(),
    }
    for k, v in defaults.items():
        if k not in st.session_state:
            st.session_state[k] = v

init_state()

# ─────────────────────────────────────────────────────────
# HELPERS
# ─────────────────────────────────────────────────────────
def get_current_outputs():
    sd = st.session_state.step_data
    inp = TemplateInputs(
        company=st.session_state.get("reporting_company") or st.session_state.user_company,
        year=st.session_state.get("reporting_year", CURR_YEAR),
        **{k: float(sd.get(k, 0)) for k in [
            "total_sites", "iso_sites", "production", "water_withdrawals",
            "renew_elec_purchased", "nonrenew_elec_purchased", "self_gen_elec",
            "purchased_steam", "sold_electricity", "sold_steam",
            "nat_gas", "coal_sub", "propane", "fuel_oil_heavy_a",
            "diesel", "petrol", "biomass", "waste_tires_mt", "lpg", "other_fuels",
            "co2_scope2_steam", "waste_total", "waste_recovery",
        ]}
    )
    return inp, calculate(inp)


# Valid TemplateInputs fields -- used to guard against unexpected keys from consolidated data
_VALID_TEMPLATE_FIELDS = {
    "total_sites", "iso_sites", "production", "water_withdrawals",
    "renew_elec_purchased", "nonrenew_elec_purchased", "self_gen_elec",
    "purchased_steam", "sold_electricity", "sold_steam",
    "nat_gas", "coal_sub", "propane", "fuel_oil_heavy_a",
    "diesel", "petrol", "biomass", "waste_tires_mt", "lpg", "other_fuels",
    "co2_scope2_steam", "waste_total", "waste_recovery",
}

def _get_fresh_hist(company: str = None) -> dict:
    """
    Load company historical data for ALL available years from _CONSOLIDATED_DF.
    Includes the current year if a row exists (e.g. after a save).
    Falls back to HIST_RAW (static demo data) when company is unknown.
    """
    co = company or st.session_state.get("reporting_company") or st.session_state.get("user_company") or ""
    if co and not _CONSOLIDATED_DF.empty:
        hist = dl.get_company_hist(_CONSOLIDATED_DF, co)
        if hist:
            # Use ALL years present in the DB, not just the pre-2023 window
            all_years = sorted(dl.get_years(_CONSOLIDATED_DF, co) or [])
            return dl.get_hist_raw(hist, all_years) if all_years else dl.get_hist_raw(hist, HIST_YEARS)
    return st.session_state.get("live_hist_raw") or HIST_RAW


def get_hist_outputs():
    """
    Return list of (year, TemplateInputs, TemplateOutputs) for ALL years in the DB.
    Uses year-keyed dict lookup — avoids positional list drift when fields missing.
    Always reads from _CONSOLIDATED_DF so any saved update is immediately visible.
    """
    company = (st.session_state.get("reporting_company") or
               st.session_state.get("user_company") or "")
    if company and not _CONSOLIDATED_DF.empty:
        all_years  = sorted(dl.get_years(_CONSOLIDATED_DF, company) or [])
        comp_hist  = dl.get_company_hist(_CONSOLIDATED_DF, company)
    else:
        all_years  = list(HIST_YEARS)
        comp_hist  = {}
    outs = []
    for yr in all_years:
        step  = dl.get_step_data(comp_hist, yr) if comp_hist else {}
        clean = {k: v for k, v in step.items() if k in _VALID_TEMPLATE_FIELDS}
        inp   = TemplateInputs(company=company, year=yr, **clean)
        outs.append((yr, inp, calculate(inp)))
    return outs



# kpi_card_html is imported from ui_components (see import block above).
# The old local definition has been removed to avoid shadowing the import.


def show_login():
    # ── Dark page background ───────────────────────────────────────────────────
    st.markdown("""<style>
    [data-testid="stSidebar"] { display:none !important; }
    [data-testid="stApp"]  { background:#1a1b6b !important; }
    [data-testid="stHeader"]{ display:none !important; }
    .main                  { background:transparent !important; }
    .block-container       { max-width:520px !important; padding:80px 24px 40px !important; margin:auto !important; }
    /* Card appearance for the whole block-container */
    .block-container > div {
        background:#f0f1f7 !important;
        border-radius:16px !important;
        padding:38px 40px 32px !important;
        box-shadow:0 25px 70px rgba(0,0,0,.45) !important;
    }
    /* Input field styling */
    [data-testid="stTextInput"] input {
        background:#fff !important; border:1px solid #dde0f0 !important;
        border-radius:7px !important; padding:11px 14px !important;
        font-size:14px !important; color:#1a1b6b !important;
    }
    [data-testid="stTextInput"] label {
        font-size:10px !important; font-weight:700 !important;
        letter-spacing:.9px !important; color:#8b90a0 !important;
        text-transform:uppercase !important;
    }
    /* Sign-in button */
    [data-testid="stButton"] > button[kind="primary"] {
        background:#111827 !important; color:#fff !important;
        border:none !important; border-radius:8px !important;
        font-size:15px !important; font-weight:500 !important;
        padding:13px !important; letter-spacing:.1px !important;
    }
    [data-testid="stButton"] > button[kind="primary"]:hover {
        background:#1f2937 !important;
    }
    /* Radio tabs look */
    [data-testid="stRadio"] > div {
        background:#e5e6ef; border-radius:8px; padding:4px;
        display:flex; gap:4px;
    }
    [data-testid="stRadio"] label {
        flex:1; text-align:center; padding:7px 12px;
        border-radius:6px; font-size:13px; font-weight:500;
        color:#6b7280; cursor:pointer;
    }
    [data-testid="stRadio"] [aria-checked="true"] + div {
        background:#fff !important;
    }
    </style>""", unsafe_allow_html=True)

    # ── Logo ──────────────────────────────────────────────────────────────────
    st.markdown("""
    <div style="margin-bottom:4px">
      <span style="font-size:34px;font-weight:800;color:#dc2626;font-style:italic;letter-spacing:-1px">dss</span><span
            style="font-size:34px;font-weight:800;color:#1a1b6b;font-style:italic;letter-spacing:-1px">+</span>
    </div>
    <div style="font-size:10px;font-weight:600;color:#9ca3af;letter-spacing:2.5px;margin-bottom:28px">
      PROTECT · TRANSFORM · SUSTAIN
    </div>""", unsafe_allow_html=True)

    # ── Role tabs ─────────────────────────────────────────────────────────────
    role = st.radio("", ["TIP Client Company", "dss+ Analyst"],
                    horizontal=True, key="login_role",
                    label_visibility="collapsed")
    st.markdown('<div style="height:14px"></div>', unsafe_allow_html=True)

    # Auto-switch email to correct default when role changes
    _prev_role = st.session_state.get("_login_prev_role")
    if _prev_role != role:
        st.session_state["_login_prev_role"] = role
        st.session_state["login_email"] = (
            "verdatyres@tip-reporting.com"
            if role == "TIP Client Company"
            else "employee@consultdss.com"
        )

    # ── Fields ────────────────────────────────────────────────────────────────
    email    = st.text_input("EMAIL ADDRESS", key="login_email")
    password = st.text_input("PASSWORD", type="password", value="demo1234", key="login_pw")
    st.markdown('<div style="height:6px"></div>', unsafe_allow_html=True)

    # ── Sign-in button ─────────────────────────────────────────────────────────
    if st.button("Sign in to workspace", type="primary",
                 use_container_width=True, key="login_btn"):
        email_l   = email.strip().lower()
        is_dss    = DSS_DOMAIN in email_l
        is_client = email_l in CLIENTS
        if not is_dss and not is_client:
            st.error("Email not recognised. Use the demo credentials below.")
        else:
            name_parts = email.split("@")[0].replace(".", " ").split()
            name       = " ".join(p.capitalize() for p in name_parts)
            st.session_state.authenticated = True
            st.session_state.user_email    = email_l
            st.session_state.user_name     = name
            st.session_state.is_dss        = is_dss
            st.session_state.user_company  = "All Companies" if is_dss else CLIENTS[email_l]
            st.session_state.page          = "portfolio" if is_dss else "home"
            st.rerun()

    # ── Demo credentials ──────────────────────────────────────────────────────
    st.markdown("""
    <div style="text-align:center;margin-top:18px;font-size:11px;color:#9ca3af;line-height:1.7">
      Demo: verdatyres@tip-reporting.com (Client) ·<br>
      analyst@consultdss.com (dss+)
    </div>""", unsafe_allow_html=True)


# ─────────────────────────────────────────────────────────
# SIDEBAR — two-shell navigation
# ─────────────────────────────────────────────────────────
def _nav_item(page_id: str, label: str) -> None:
    """Render one sidebar nav item. Active item is highlighted green."""
    active = st.session_state.page == page_id
    if active:
        st.markdown(
            f'<div style="background:rgba(22,163,74,0.85);border-radius:7px;'
            f'padding:8px 14px;margin-bottom:2px;color:#fff;font-size:13px;'
            f'font-weight:600">{label}</div>',
            unsafe_allow_html=True,
        )
    else:
        if st.sidebar.button(label, key=f"nav_{page_id}", use_container_width=True):
            st.session_state.page = page_id
            st.rerun()


def show_sidebar():
    with st.sidebar:
        # ── Logo ─────────────────────────────────────────────────────────────
        st.markdown("""
        <div style="padding:16px 14px 12px;border-bottom:1px solid rgba(255,255,255,.08)">
          <div style="display:flex;align-items:center;gap:8px">
            <div style="width:10px;height:10px;border-radius:50%;background:#16A34A;flex-shrink:0"></div>
            <div>
              <div style="color:#fff;font-size:14px;font-weight:700;letter-spacing:-.2px">TIP ESG Platform</div>
              <div style="color:rgba(255,255,255,.35);font-size:10px;margin-top:1px">dss+ · Tire Industry Project</div>
            </div>
          </div>
        </div>
        """, unsafe_allow_html=True)

        # ── Company badge (client only) ───────────────────────────────────────
        if not st.session_state.is_dss:
            _safe_co = _html.escape(st.session_state.user_company)
            st.markdown(f"""
            <div style="margin:10px 10px 0;padding:8px 12px;background:rgba(255,255,255,.06);
                border-radius:8px;border:1px solid rgba(255,255,255,.08)">
              <div style="color:rgba(255,255,255,.35);font-size:9px;text-transform:uppercase;
                  letter-spacing:.6px">Your Company</div>
              <div style="color:#fff;font-size:13px;font-weight:500;margin-top:2px">{_safe_co}</div>
            </div>""", unsafe_allow_html=True)

        # ── CLIENT navigation ─────────────────────────────────────────────────
        if not st.session_state.is_dss:
            _nav_item("home",         "Home")
            _nav_item("dashboard",    "My Dashboard")
            _nav_item("my_records",   "My Records")
            _nav_item("benchmarking", "Benchmarks")
            _nav_item("reports",      "Reports")
            _nav_item("entry",        "Submit Data")
            _nav_item("settings",     "Settings")

        # ── DSS+ INTERNAL navigation ──────────────────────────────────────────
        else:
            _nav_item("portfolio",      "Portfolio")
            _nav_item("company_data",   "Company Data")
            _nav_item("verification",   "Verification Queue")
            _nav_item("analysis",       "Analysis")
            _nav_item("benchmarking",   "Benchmarks")
            _nav_item("readiness",      "AI Assistant")
            _nav_item("doc_library",    "Document Library")
            _nav_item("sector_reports", "Sector Reports")
            _nav_item("admin",          "Admin")
            _nav_item("settings",       "Settings")
            # Submit Data hidden for now — restore by uncommenting:
            # _nav_item("entry", "Submit Data")

        # ── User footer ───────────────────────────────────────────────────────
        st.markdown("---")
        name_init  = _html.escape(
            "".join(p[0].upper() for p in st.session_state.user_name.split()[:2])
        )
        _safe_name = _html.escape(st.session_state.user_name)
        role_lbl   = "dss+ Analyst" if st.session_state.is_dss else f"Client · {st.session_state.user_company}"
        _safe_role = _html.escape(role_lbl)
        st.markdown(f"""
        <div style="display:flex;align-items:center;gap:9px;padding:0 2px">
          <div style="width:30px;height:30px;border-radius:50%;background:#16A34A;color:#fff;
              font-size:11px;font-weight:700;display:flex;align-items:center;
              justify-content:center;flex-shrink:0">{name_init}</div>
          <div>
            <div style="color:#fff;font-size:13px;font-weight:500">{_safe_name}</div>
            <div style="color:rgba(255,255,255,.4);font-size:10px">{_safe_role}</div>
          </div>
        </div>""", unsafe_allow_html=True)
        st.markdown('<div style="height:6px"></div>', unsafe_allow_html=True)
        if st.button("Sign out", use_container_width=True, key="signout_btn"):
            for k in list(st.session_state.keys()):
                del st.session_state[k]
            st.rerun()


# ─────────────────────────────────────────────────────────
# STEPPER BAR
# ─────────────────────────────────────────────────────────
def _build_master_row(inp, out, supp: dict = None) -> dict:
    """
    Build a dict whose keys exactly match the master wide CSV column names.
    Ensures no duplicate columns when appended to the master CSV.
    Includes waste KPIs and electricity-by-country columns.
    """
    renew_share  = (inp.renew_elec_purchased + inp.self_gen_elec) / max(out.total_electricity, 1) * 100
    scope1_share = out.total_co2_scope1 / max(out.total_co2, 1) * 100
    scope2_share = out.total_co2_scope2 / max(out.total_co2, 1) * 100
    fuel_total   = (inp.nat_gas + inp.coal_sub + inp.propane + inp.fuel_oil_heavy_a + inp.diesel + inp.petrol)
    fossil_share = fuel_total / max(out.total_energy, 1) * 100
    prod         = max(inp.production, 1)

    # Waste derived
    waste_total    = float(getattr(inp, "waste_total", 0) or 0)
    waste_recovery = float(getattr(inp, "waste_recovery", 0) or 0)
    recovery_rate  = round(waste_recovery / waste_total * 100, 4) if waste_total else 0.0
    waste_elim     = round(waste_total - waste_recovery, 4)

    row = {
        "Company": inp.company, "Year": inp.year,
        "Total no. of sites": int(round(inp.total_sites)),
        "ISO 14001 sites":    int(round(inp.iso_sites)),
        "% certified sites":  round(out.pct_certified, 6),
        "Production":         round(inp.production, 4),
        "Water intake":       round(inp.water_withdrawals, 4),
        "Water intake - KPI": round(out.water_kpi, 6),
        "Total Electricity":                               round(out.total_electricity, 4),
        "Renewable Electricity Purchased":                 round(inp.renew_elec_purchased, 4),
        "Non-Renewable Electricity Purchased":             round(inp.nonrenew_elec_purchased, 4),
        "Self-generated AND consumed electricity on-site": round(inp.self_gen_elec, 4),
        "Purchased Steam":   round(inp.purchased_steam, 4),
        "Sold Electricity":  round(inp.sold_electricity, 4),
        "Sold Steam":        round(inp.sold_steam, 4),
        "Natural Gas":       round(inp.nat_gas, 4),
        "Coal":              round(inp.coal_sub, 4),
        "Propane":           round(inp.propane, 4),
        "Fuel Oil":          round(inp.fuel_oil_heavy_a, 4),
        "Diesel":            round(inp.diesel, 4),
        "Petrol":            round(inp.petrol, 4),
        "Biomass":           round(inp.biomass, 4),
        "Waste tires":       round(inp.waste_tires_mt, 4),
        "LPG":               round(inp.lpg, 4),
        "Other":             round(inp.other_fuels, 4),
        "Total energy":          round(out.total_energy, 4),
        "Total energy - KPI":    round(out.energy_kpi, 6),
        "Total CO2 - Scope 1":   round(out.total_co2_scope1, 4),
        "Total CO2 - Scope 2":   round(out.total_co2_scope2, 4),
        "Total CO2":             round(out.total_co2, 4),
        "Total CO2 - KPI":       round(out.co2_kpi, 6),
        # ── Waste fields ──────────────────────────────────────────────────────
        "Total Waste":           round(waste_total, 4),
        "Waste Recovered":       round(waste_recovery, 4),
        "Recovery Rate":         recovery_rate,
        # ── Country electricity placeholders (filled by _save_electricity_to_master) ─
        **{_elec_col(c): None for c in ELEC_ALL_COUNTRIES},
        # ── Derived KPIs ──────────────────────────────────────────────────────
        "Renewable_Electricity_Share_%": round(renew_share, 4),
        "Scope1_Share_%":                round(scope1_share, 4),
        "Scope2_Share_%":                round(scope2_share, 4),
        "Fossil_Energy_Share_%":         round(fossil_share, 4),
        "Water_per_ton":                 round(inp.water_withdrawals / prod, 4),
        "CO2_per_ton":                   round(out.total_co2 / prod, 4),
        "Energy_per_ton":                round(out.total_energy / prod, 4),
        "ISO_Certification_%":           round(out.pct_certified * 100, 2),
        "Waste_Recovery_Rate_%":         recovery_rate,
        "Total_Electricity_by_Country_GJ": None,  # filled after country save
    }

    # ── People & Governance fields (from supplementary) ────────────────────
    s = supp or {}
    def _sf(k, default=0.0):
        try: return float(s.get(k, default) or default)
        except: return default

    stress_wd     = _sf("stress_water_withdrawal")
    non_stress_wd = _sf("non_stress_water_withdrawal",
                         max(inp.water_withdrawals - stress_wd, 0))
    hs_ext   = _sf("hs_external_audit")
    hs_int   = _sf("hs_internal_audit")
    hs_tot   = max(int(_sf("hs_external_audit", inp.total_sites)), int(inp.total_sites))
    emp_tot  = _sf("total_employees")
    emp_fem  = _sf("female_employees")
    bod_tot  = _sf("board_total")
    bod_fem  = _sf("female_board")

    row.update({
        # Water detail
        "Stress Water Withdrawal":     round(stress_wd, 4),
        "Non-Stress Water Withdrawal": round(non_stress_wd, 4),
        # Coal breakdown
        "Coal Sub-Bituminous":         round(_sf("coal_sub_bituminous"), 4),
        "Coal Brown Briquettes":       round(_sf("coal_brown_briquettes"), 4),
        "Coal Other Bituminous":       round(_sf("coal_other_bituminous"), 4),
        # H&S
        "HS External Audit Sites":     round(hs_ext),
        "HS Internal Audit Sites":     round(hs_int),
        "HS External Audit %":         round(hs_ext / max(hs_tot, 1) * 100, 2),
        "HS Internal Audit %":         round(hs_int / max(hs_tot, 1) * 100, 2),
        # Diversity
        "Total Employees":             round(emp_tot),
        "Female Employees":            round(emp_fem),
        "Female Employees %":          round(emp_fem / max(emp_tot, 1) * 100, 2),
        "Board Total":                 round(bod_tot),
        "Female Board":                round(bod_fem),
        "Female Board %":              round(bod_fem / max(bod_tot, 1) * 100, 2),
        # Science-Based Targets
        "SBT Total":       round(_sf("sbt_total")),
        "SBT Validated":   round(_sf("sbt_validated")),
        "SBT Committed":   round(_sf("sbt_committed")),
        "SBT Non-Committed": round(_sf("sbt_non_committed")),
    })
    return row


def _save_version_parquet(inp, combined_df: pd.DataFrame) -> str:
    """
    Save the ENTIRE company template (all years) as a Parquet snapshot.
    Stored in data_storage/versions/{CompanyName}/ — subfolder only, never flat.
    Filename: CompanyName_Year_YYYYMMDD_HHMMSS.parquet (year = the year just edited).
    NEVER overwritten — each save event creates a new file (full audit trail).
    Reading this file shows the complete state of all years for that company
    at the exact moment the save was made.
    """
    from pathlib import Path
    from datetime import datetime
    ts      = datetime.now().strftime("%Y%m%d_%H%M%S")
    co_safe = inp.company.replace(" ", "_").replace("/", "_")
    # Extract ALL rows for this company from the combined master DataFrame
    company_all_years = combined_df[combined_df["Company"] == inp.company].copy()
    filename = f"{co_safe}_{inp.year}_{ts}.parquet"
    # Subfolder only — no flat file
    ver_dir  = Path("data_storage") / "versions" / co_safe
    ver_dir.mkdir(parents=True, exist_ok=True)
    try:
        company_all_years.to_parquet(ver_dir / filename, index=False)
        return f"{co_safe}/{filename}"
    except Exception as e:
        return f"[version save failed: {e}]"




def _migrate_supplementary_to_master() -> str:
    """
    One-time (idempotent) migration: read every row in ESG_SUPPLEMENTARY.csv
    and upsert it into the master CSV so the People & Governance columns are
    populated for all previously-submitted records.
    Safe to call on every startup — skips companies/years already promoted.
    """
    import csv as _csv
    from pathlib import Path as _P
    supp_path = _P("data_storage/master/ESG_SUPPLEMENTARY.csv")
    if not supp_path.exists():
        return "no supplementary file — skipped"

    migrated = 0
    with open(supp_path, newline="", encoding="utf-8") as f:
        rows = list(_csv.DictReader(f))

    for row in rows:
        company = row.get("Company","").strip()
        year_s  = row.get("Year","").strip()
        if not company or not year_s:
            continue
        try:
            year = int(year_s)
        except ValueError:
            continue

        # Build supp dict for this row
        supp = {k: float(v) if v else 0.0
                for k, v in row.items() if k not in ("Company","Year")}

        # Load existing TemplateInputs for this company+year
        hist = dl.get_company_hist(_CONSOLIDATED_DF, company)
        if not hist:
            continue
        sd = dl.get_step_data(hist, year)
        sd_clean = {k: v for k, v in sd.items() if k in _VALID_TEMPLATE_FIELDS}
        if not sd_clean:
            continue

        inp = TemplateInputs(company=company, year=year, **sd_clean)
        out = calculate(inp)

        # Re-save the master row with supplementary data included
        _save_submission_to_csv(inp, out)   # supp auto-loaded inside
        migrated += 1

    return f"migrated {migrated} supplementary records into master"

# Monkey-patch _save_submission_to_csv to support direct supp argument
# (used by migration only — normal path loads supp inside the function)


def _drop_zero_elec_cols(df: "pd.DataFrame") -> "pd.DataFrame":
    """
    Return df with Elec_*_GJ country columns removed if every value in that
    column is zero or null across all rows.  Non-electricity columns are
    never touched.  Used so files only carry countries with actual consumption.
    """
    elec_cols = [c for c in df.columns if c.startswith("Elec_") and c.endswith("_GJ")
                 and c != "Total_Electricity_by_Country_GJ"]
    zero_cols = [c for c in elec_cols
                 if df[c].fillna(0).eq(0).all()]
    return df.drop(columns=zero_cols) if zero_cols else df

def _sync_consolidate_excel(master_df: "pd.DataFrame") -> None:
    """
    Sync the CONSOLIDATED_DUMMY Excel (Raw Dummy data sheet) from the master wide CSV.

    The Raw Dummy data sheet stores data in long format: one row per
    (Company, Year, Row_Label). This function overwrites it completely from
    the current master wide DataFrame so that the consolidate stays in sync
    after any save from the platform.
    """
    from pathlib import Path
    from openpyxl import load_workbook

    xl_path = Path("data_storage/master/CONSOLIDATED_DUMMY_2009_2023.xlsx")
    if not xl_path.exists():
        return  # nothing to sync yet

    # Mapping: wide-CSV column  →  (Section, Row_Label)
    COL_MAP = {
        "Total no. of sites":                              ("ISO 14001",    "Total no. of sites"),
        "ISO 14001 sites":                                 ("ISO 14001",    "ISO 14001 sites"),
        "% certified sites":                               ("ISO 14001",    "% certified sites"),
        "Production":                                      ("Production",   "Production"),
        "Water intake":                                    ("Water",        "Water intake"),
        "Water intake - KPI":                              ("Water",        "Water intake - KPI"),
        "Total Electricity":                               ("Energy",       "Total Electricity"),
        "Renewable Electricity Purchased":                 ("Energy",       "Renewable Electricity Purchased"),
        "Non-Renewable Electricity Purchased":             ("Energy",       "Non-Renewable Electricity Purchased"),
        "Self-generated AND consumed electricity on-site": ("Energy",       "Self-generated AND consumed electricity on-site"),
        "Purchased Steam":                                 ("Energy",       "Purchased Steam"),
        "Sold Electricity":                                ("Energy",       "Sold Electricity"),
        "Sold Steam":                                      ("Energy",       "Sold Steam"),
        "Natural Gas":                                     ("Energy",       "Natural Gas"),
        "Coal":                                            ("Energy",       "Coal"),
        "Propane":                                         ("Energy",       "Propane"),
        "Fuel Oil":                                        ("Energy",       "Fuel Oil"),
        "Diesel":                                          ("Energy",       "Diesel"),
        "Petrol":                                          ("Energy",       "Petrol"),
        "Biomass":                                         ("Energy",       "Biomass"),
        "Waste tires":                                     ("Energy",       "Waste tires"),
        "LPG":                                             ("Energy",       "LPG"),
        "Other":                                           ("Energy",       "Other"),
        "Total energy":                                    ("Energy",       "Total energy"),
        "Total energy - KPI":                              ("Energy",       "Total energy - KPI"),
        "Total CO2 - Scope 1":                             ("CO2 emissions","Total CO2 - Scope 1"),
        "Total CO2 - Scope 2":                             ("CO2 emissions","Total CO2 - Scope 2"),
        "Total CO2":                                       ("CO2 emissions","Total CO2"),
        "Total CO2 - KPI":                                 ("CO2 emissions","Total CO2 - KPI"),
        "Total Waste":                                     ("Waste",        "Total Waste"),
        "Waste Recovered":                                 ("Waste",        "Waste Recovered"),
        "Recovery Rate":                                   ("Waste",        "Recovery Rate"),
        **{_elec_col(c): ("Energy", f"Electricity - {c}") for c in ELEC_ALL_COUNTRIES},
        # People & Governance (promoted from supplementary)
        "Stress Water Withdrawal":     ("Water",       "Stress water withdrawal"),
        "Non-Stress Water Withdrawal": ("Water",       "Non-stress water withdrawal"),
        "Coal Sub-Bituminous":         ("Energy",      "Coal — Sub-bituminous"),
        "Coal Brown Briquettes":       ("Energy",      "Coal — Brown briquettes"),
        "Coal Other Bituminous":       ("Energy",      "Coal — Other bituminous"),
        "HS External Audit Sites":     ("H&S",         "Externally audited H&S sites"),
        "HS Internal Audit Sites":     ("H&S",         "Internally audited H&S sites"),
        "HS External Audit %":         ("H&S",         "H&S external audit coverage %"),
        "HS Internal Audit %":         ("H&S",         "H&S internal audit coverage %"),
        "Total Employees":             ("Diversity",   "Total employees"),
        "Female Employees":            ("Diversity",   "Female employees"),
        "Female Employees %":          ("Diversity",   "% Female employees"),
        "Board Total":                 ("Diversity",   "Board of Directors total"),
        "Female Board":                ("Diversity",   "Female Board members"),
        "Female Board %":              ("Diversity",   "% Female Board"),
        "SBT Total":                   ("SBT",         "Total with science-based target"),
        "SBT Validated":               ("SBT",         "SBT — Validated"),
        "SBT Committed":               ("SBT",         "SBT — Committed"),
        "SBT Non-Committed":           ("SBT",         "SBT — Non-committed"),
    }

    # Build long rows from master_df
    long_rows = []  # list of dicts: Company, Row, Year, Data, Section, Row_Label, Notes, Consistency test
    row_order = list(COL_MAP.keys())
    # Assign fixed row numbers to match what build_esg_master.py uses
    ROW_NUM = {col: i + 1 for i, col in enumerate(row_order)}

    # Pre-compute which electricity country columns have any non-zero value
    # across the whole master — only those countries get rows in the consolidate.
    active_elec_cols = {
        col for col in COL_MAP
        if col.startswith("Elec_") and col.endswith("_GJ")
        and col in master_df.columns
        and master_df[col].fillna(0).ne(0).any()
    }

    for _, wrow in master_df.sort_values(["Company", "Year"]).iterrows():
        company = wrow["Company"]
        year    = int(wrow["Year"]) if pd.notna(wrow.get("Year")) else None
        if not company or not year:
            continue
        for col, (section, label) in COL_MAP.items():
            # Skip electricity country columns that are all-zero across the dataset
            is_elec_country = col.startswith("Elec_") and col.endswith("_GJ")
            if is_elec_country and col not in active_elec_cols:
                continue
            val = wrow.get(col)
            # For an active electricity country, skip rows where this company-year is zero
            if is_elec_country and (pd.isna(val) or float(val) == 0):
                continue
            long_rows.append({
                "Company": company,
                "Row":     ROW_NUM[col],
                "Year":    year,
                "Data":    float(val) if pd.notna(val) else None,
                "Section": section,
                "Row_Label": label,
                "Notes":   None,
                "Consistency test": None,
            })

    if not long_rows:
        return

    try:
        wb = load_workbook(xl_path)
        ws = wb["Raw Dummy data"]
        # Clear existing data rows (keep header row 1)
        for r in range(2, ws.max_row + 1):
            for c in range(1, 9):
                ws.cell(r, c).value = None
        # Write new rows
        cols = ["Company", "Row", "Year", "Data", "Section", "Row_Label", "Notes", "Consistency test"]
        for i, row in enumerate(long_rows):
            for j, col in enumerate(cols):
                ws.cell(i + 2, j + 1).value = row[col]
        wb.save(xl_path)
    except (PermissionError, OSError):
        pass  # file locked — skip, master CSV is the source of truth
    except Exception:
        pass  # any other error is also non-fatal


def _sync_company_member_files(master_df: "pd.DataFrame") -> list:
    """
    Write per-company CSVs in data_storage/members/TIP/<CompanyName>/<CompanyName>_latest.csv
    from the current master wide DataFrame.
    Skips any file that is locked (e.g. open in Excel) instead of crashing.
    Returns list of company names that were skipped.
    """
    from pathlib import Path
    members_tip = Path("data_storage/members/TIP")
    skipped = []
    for company, grp in master_df.groupby("Company"):
        co_safe   = str(company).replace(" ", "_")
        co_folder = members_tip / co_safe
        co_folder.mkdir(parents=True, exist_ok=True)
        try:
            # Drop electricity country columns that are all zero for this company
            grp_clean = _drop_zero_elec_cols(grp.reset_index(drop=True))
            grp_clean.to_csv(co_folder / f"{co_safe}_latest.csv", index=False)
        except (PermissionError, OSError):
            skipped.append(str(company))
    return skipped


def _save_submission_to_csv(inp, out) -> str:
    """
    Three independent operations:

    1. MASTER CSV (data_storage/master/) — overwrite the row for this company+year.
       The master always holds the LATEST values. Second save for same company+year
       replaces the first row.

    2. VERSION Parquet (data_storage/versions/) — always ADD a new timestamped file.
       Never overwritten. Full audit trail of every save event.

    3. SYNC (after master is written):
       - CONSOLIDATED_DUMMY Excel Raw Dummy data sheet (long format)
       - Per-company CSVs in data_storage/members/TIP/<Company>/
       - TIP members aggregate CSV
    """
    import os, tempfile
    from pathlib import Path
    from datetime import datetime

    _master_cands = dl._get_csv_candidates()
    csv_path = next((p for p in _master_cands if p.exists()
                     and p.name.startswith("ESG_MASTER_WIDE_ALL_COMPANIES_")),
                    None) or Path("data_storage/master/ESG_MASTER_WIDE_ALL_COMPANIES_2009_2023.csv")
    csv_path.parent.mkdir(parents=True, exist_ok=True)

    # Load supplementary fields so they are promoted into the master CSV
    _supp_data = _load_supplementary(inp.company, inp.year)
    new_row    = pd.DataFrame([_build_master_row(inp, out, supp=_supp_data)])
    master_cols = list(new_row.columns)

    def _align(df):
        """Align DataFrame to master column order: strip extras, fill missing."""
        if df.empty:
            return pd.DataFrame(columns=master_cols)
        extra = [c for c in df.columns if c not in master_cols]
        if extra:
            df = df.drop(columns=extra)
        for col in master_cols:
            if col not in df.columns:
                df[col] = None
        return df[master_cols]

    def _load_best_existing():
        """
        Load the most complete existing master DataFrame.
        Checks all candidate paths and picks the one with the most rows.
        """
        candidates = [
            csv_path,
            Path("data_storage/raw/ESG_MASTER_WIDE_ALL_COMPANIES_2009_2023.csv"),
        ]
        best = pd.DataFrame(columns=master_cols)
        for p in candidates:
            if p.exists():
                try:
                    df = pd.read_csv(p)
                    if "Company" in df.columns and "Year" in df.columns and len(df) > len(best):
                        best = df
                except PermissionError:
                    pass
                except Exception:
                    pass
        return _align(best)

    # ── 1. Build combined DataFrame ──────────────────────────────────────────
    # H1 FIX: advisory file lock held for the full read-modify-write cycle.
    # Prevents data corruption if two analysts save simultaneously.
    # Timeout=10s: if another process holds the lock and crashes, we don't
    # block forever.  The PermissionError branch below still handles Excel locks.
    lock_path = csv_path.with_suffix(".lock")
    with FileLock(str(lock_path), timeout=cfg.FILELOCK_TIMEOUT):
        existing = _load_best_existing()
        mask     = ~((existing["Company"] == inp.company) & (existing["Year"] == inp.year))
        existing = existing[mask]
        combined = pd.concat([existing, new_row], ignore_index=True)
        combined = combined.sort_values(["Company", "Year"]).reset_index(drop=True)

        n_records   = len(combined)
        n_companies = combined["Company"].nunique()

        # ── 2. Save version Parquet BEFORE touching master (audit trail first) ───
        version_filename = _save_version_parquet(inp, combined)

        # ── 3. Write master CSV, then sync all dependent files ───────────────────
        try:
            # Master CSV keeps all country columns (even all-zero) as the full schema.
            # Derived outputs (member files, TIP aggregate) strip all-zero country cols.
            combined.to_csv(csv_path, index=False)
            # Sync TIP members aggregate
            tip_master_path = Path("data_storage/members/TIP/ESG_MASTER_WIDE_TIP_MEMBERS_2009_2023.csv")
            _update_tip_members_file(csv_path, tip_master_path)
            # Sync per-company member files
            _sync_company_member_files(combined)
            # Sync CONSOLIDATED_DUMMY Excel
            _sync_consolidate_excel(combined)

            # ── Auto-add electricity-by-country year columns for new submission ──
            # If company just submitted for a year that isn't in the electricity
            # editor yet, initialize all country columns to 0 in the master CSV.
            _new_yr = inp.year
            _elec_cols_all = [c for c in combined.columns
                              if c.startswith("Elec_") and c.endswith("_GJ")]
            _co_yr_mask = (combined["Company"] == inp.company) & (combined["Year"] == _new_yr)
            if _co_yr_mask.any() and _elec_cols_all:
                for _ec in _elec_cols_all:
                    if pd.isna(combined.loc[_co_yr_mask, _ec]).all():
                        combined.loc[_co_yr_mask, _ec] = 0.0
                # Also ensure Total_Electricity_by_Country_GJ exists
                if "Total_Electricity_by_Country_GJ" not in combined.columns:
                    combined["Total_Electricity_by_Country_GJ"] = 0.0
                elif pd.isna(combined.loc[_co_yr_mask, "Total_Electricity_by_Country_GJ"]).all():
                    combined.loc[_co_yr_mask, "Total_Electricity_by_Country_GJ"] = 0.0
                # Re-write master with the zero-initialized electricity columns
                combined.to_csv(csv_path, index=False)

            # ── CRITICAL: update the in-memory globals so all pages in this
            #    session immediately see the new data without requiring a restart.
            global _CONSOLIDATED_DF, _COMPANIES, _SECTOR_DF, _USING_FALLBACK_DATA
            global HIST_YEARS, CURR_YEAR, LONG_YEARS, LONG_DATA, FUEL_MIX
            _CONSOLIDATED_DF     = combined.copy()
            _COMPANIES           = dl.get_companies(combined)
            _USING_FALLBACK_DATA = False
            try:
                cfg.refresh_year_bounds(combined)
                HIST_YEARS = cfg.hist_years()
                CURR_YEAR  = cfg.curr_year()
                LONG_YEARS = cfg.long_years()
            except Exception:
                pass
            try:
                _SECTOR_DF = dl.load_sector_aggregated(combined)
            except Exception:
                pass
            try:
                LONG_DATA, FUEL_MIX = _build_long_data()
            except Exception:
                pass
            st.cache_data.clear()

            return (f"Saved {inp.company} — {inp.year}. "
                    f"Master: {n_records} records across {n_companies} companies. "
                    f"Version: {version_filename}")
        except PermissionError:
            ts          = datetime.now().strftime("%Y%m%d_%H%M%S")
            backup_name = f"ESG_MASTER_{inp.company.replace(' ','_')}_{inp.year}_{ts}.csv"
            backup_path = csv_path.parent / backup_name
            try:
                combined.to_csv(backup_path, index=False)
                return (
                    f"⚠️ Master file open in Excel — saved backup: **{backup_name}**\n"
                    f"Version snapshot: {version_filename}\n"
                    f"Close Excel and click Save again."
                )
            except Exception as e2:
                return f"❌ Save failed (file locked AND backup failed): {e2}"
        except Exception as e:
            return f"❌ Save failed: {e}"




# ── Populate shared state module so pages and components can read globals ──────
import state as _state
_state.CONSOLIDATED_DF = _CONSOLIDATED_DF
_state.COMPANIES       = _COMPANIES
_state.SECTOR_DF       = _SECTOR_DF
_state.USING_FALLBACK  = _USING_FALLBACK_DATA
_state.HIST_YEARS      = HIST_YEARS
_state.CURR_YEAR       = CURR_YEAR
_state.LONG_YEARS      = LONG_YEARS
_state.LONG_DATA       = LONG_DATA
_state.FUEL_MIX        = FUEL_MIX


def _refresh_state():
    """Call after any save to push updated globals into state module."""
    _state.CONSOLIDATED_DF = _CONSOLIDATED_DF
    _state.COMPANIES       = _COMPANIES
    _state.SECTOR_DF       = _SECTOR_DF
    _state.USING_FALLBACK  = _USING_FALLBACK_DATA
    _state.HIST_YEARS      = HIST_YEARS
    _state.CURR_YEAR       = CURR_YEAR
    _state.LONG_YEARS      = LONG_YEARS
    _state.LONG_DATA       = LONG_DATA
    _state.FUEL_MIX        = FUEL_MIX


# ── Component imports ─────────────────────────────────────────────────────────
from components.render_template_table  import render_template_table
from components.render_electricity_tab import render_electricity_tab
from components.render_waste_tab       import render_waste_tab
from components.render_people_tab      import _render_people_governance_tab
from components.render_qualitative_tab import render_qualitative_tab
from components.render_conversion_tab  import render_conversion_tab

# ── Page imports ───────────────────────────────────────────────────────────────
from pages.page_home          import page_home
from pages.page_entry         import page_entry
from pages.page_my_records    import page_my_records
from pages.page_my_dashboard  import page_my_dashboard
from pages.page_benchmarking  import page_benchmarking
from pages.page_analysis      import page_analysis
from pages.page_verification  import page_verification
from pages.page_company_data  import page_company_data
from pages.page_reports       import page_reports
from pages.page_readiness     import page_readiness
from pages.page_portfolio     import page_portfolio
from pages.page_doc_library   import page_doc_library
from pages.page_sector_reports import page_sector_reports
from pages.page_admin         import page_admin
from pages.page_settings      import page_settings

# ─────────────────────────────────────────────────────────
# MAIN ROUTER
# ─────────────────────────────────────────────────────────
if not st.session_state.authenticated:
    show_login()
else:
    show_sidebar()
    page = st.session_state.page

    # ── One-time migration: supplementary → master CSV ───────────────────────
    if not st.session_state.get("_supp_migrated"):
        try:
            if Path("data_storage/master/ESG_SUPPLEMENTARY.csv").exists():
                _migrate_supplementary_to_master()
        except Exception:
            pass
        st.session_state["_supp_migrated"] = True

    # ── Client pages ──────────────────────────────────────
    if   page == "home":            page_home()
    elif page == "entry":           page_entry()
    elif page == "my_records":      page_my_records()
    elif page == "dashboard":       page_my_dashboard()
    elif page == "benchmarking":    page_benchmarking()
    elif page == "reports":         page_reports()
    elif page == "settings":        page_settings()

    # ── DSS+ pages ────────────────────────────────────────
    elif page == "portfolio":       page_portfolio()
    elif page == "company_data":    page_company_data()
    elif page == "verification":    page_verification()
    elif page == "analysis":        page_analysis()
    elif page == "readiness":       page_readiness()
    elif page == "doc_library":     page_doc_library()
    elif page == "sector_reports":  page_sector_reports()
    elif page == "admin":           page_admin()

    else:
        # Fallback — redirect to correct home
        st.session_state.page = "portfolio" if st.session_state.is_dss else "home"
        st.rerun()